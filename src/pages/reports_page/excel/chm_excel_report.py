import os
import math

from base_logger import logger

from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.utils import get_column_letter

from modules.utils.file_utils import get_path_from_json
from modules.utils.logic_utils import is_float

from modules.models.excel_model import ExcelModel, split_sentence_by_words, significant_figures_convert
'''
    Sₒ: standard deviation at zero analyze concentration; method detection limit is generally considered to be 3x So value
    Recovery: The percentage recovery of efficiency of isolating/purifying a substance.
    Lab Blank: A sample without the target substance used to account for background interference.
'''

#FIXME: more data is showing up since can enter in information and it will add it to the tests

class ChmExcelReport(ExcelModel):

    def __init__(self, client_info, jobNum, authors, side_comments, extra_comments, comment, sample_names, sample_data, test_info, units, recovery, so, lower_limits, upper_limits, hidden_rows, report_type=2):
        super().__init__(jobNum, report_type)
        self.authors = authors
        self.client_info = client_info
        self.comment = comment
        self.sample_names = sample_names
        self.sample_data = sample_data
        self.test_info = test_info
        self.units = units
        self.recovery = recovery
        self.lower_limits = lower_limits
        self.upper_limits = upper_limits #FIXME: currently set to None
        self.so = so

        self.side_comments = side_comments
        self.extra_comments = extra_comments

        self.extra_comment_rows = []

        self.hidden_rows = hidden_rows

        self.setup()

    def __repr__(self):
        # Provide a preview of the dictionary contents for readability
        client_info_preview = {k: v for i, (k, v) in enumerate(self.client_info.items()) if i < 3}
        sample_names_preview = {k: v for i, (k, v) in enumerate(self.sample_names.items()) if i < 3}
        sample_data_preview = {k: v for i, (k, v) in enumerate(self.sample_data.items()) if i < 3}

        return (
            f"ChmExcelReport("
            f"client_info={client_info_preview}, "
            f"jobNum={self.jobNum}, "
            f"authors={self.authors}, "
            f"comment='{self.comment}', "
            f"sample_names={sample_names_preview}, "
            f"sample_data={sample_data_preview}, "
            f"test_info={self.test_info}, "
            f"units={self.units}, "
            f"recovery={self.recovery}, "
            f"report_type={self.report_type})"
        )

    def setup(self):
        logger.info('Entering setup for ChmExcelReport')

        self.total_samples = len(self.sample_data)

        #set the total_cols
        self.total_cols = 10
        self.page_size = 78

        self.excel_page_setup()
        self.set_headers_and_footers()
        self.format_rows(self.total_samples)

        #note can insert
        self.insert_client_info(self.client_info, 'D')

    def excel_page_setup(self):
        self.ws.sheet_view.view = "pageLayout"

        # Ensure that the scaling happens automatically

        self.ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=False)

        # set custom page margins (narrow page margins)
        self.ws.page_margins = PageMargins(
            left=0.7,
            right=0.7,
            top=0.75,
            bottom=0.75,
            header=0.3,
            footer=0.3
        )

        # set the excel page settings
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0  # Automatic height fitting
        self.ws.page_setup.orientation = self.ws.ORIENTATION_PORTRAIT

        # set column dimensions
        self.ws.column_dimensions['A'].width = 21 #126 px
        self.ws.column_dimensions['J'].width = 21 #1 width is equal to 6 pixels
        self.ws.print_title_rows = '1:8' # the first two row

        for col in range(1, self.total_cols -1):
            column_letter = get_column_letter(col + 1)

            self.ws.column_dimensions[column_letter].width = 9.2 # 57 px

    def create_report(self):
        logger.info('Entering create_report')

        sample_sections, sample_placement = self.format_sample_header_names(self.sample_names)

        logger.debug(f'sample_sections: {sample_sections}')
        logger.debug(f'sample_placement: {sample_placement}')
        logger.debug(f'test_info: {self.test_info}')

        total_hidden_rows = sum(list(self.hidden_rows.values()))

        total_tests = len(self.test_info) -  total_hidden_rows
        total_tables = len(sample_placement)
        table_size = 6 + total_tests

        logger.debug(f'total_test before: {len(self.test_info)}')
        logger.debug(f'total_test after: {total_tests}')

        comment_size = len(self.comment)
        author_size = 6
        header_size = 8

        allocated_bottom_space = comment_size + author_size

        page_location = 9
        used_samples = 0
        current_page = 0

        for sample_batch, samples in enumerate(sample_placement):

            samples_amount = len(samples)
            section_names = sample_sections[sample_batch]

            logger.debug(f'sample_batch: {sample_batch}, samples_amount: {samples_amount}, samples: {samples}')
            logger.debug(f'page_location: {page_location}')

            remaining_page_size = ((current_page+1 * self.page_size) - (header_size * current_page))

            if(sample_batch+1 == total_tables):
                remaining_page_size = ((current_page+1 * self.page_size) - (header_size * current_page) - (allocated_bottom_space))

            if(page_location + table_size > remaining_page_size):
                self.insert_next_page_comment(page_location)

                current_page += 1
                next_page_start = (self.page_size * current_page) - (header_size * (current_page-1)) + 1
                page_location = next_page_start

                logger.debug(f'next_page_start: {next_page_start}')

            page_location = self.insert_sample_name(page_location, section_names)
            page_location = self.insert_test_titles(page_location, samples_amount, used_samples)
            page_location = self.insert_tests_info(page_location, samples, total_tests, total_hidden_rows)

            used_samples += samples_amount

        self.insert_report_comments(page_location, current_page, header_size, author_size)

        export_path = get_path_from_json('default_paths.json', 'reportsPath')

        file_name = f'W{self.jobNum}_chm.xlsx'
        file_path = os.path.join(export_path, file_name)

        file_path = self.save_excel(file_path)

        return file_path, file_name

    def determine_extra_comments(self):

        extra_comments = []
        total_extra_comment_rows = 1

        for row in self.extra_comment_rows:

            test_name = self.test_info[row]

            full_comment = f'{test_name}: {self.extra_comments[row]}'

            if(len(full_comment) > self.row_char_limit):
                wrapped_lines = split_sentence_by_words(full_comment)
                logger.info(f'wrapped_lines ({len(wrapped_lines)}): {wrapped_lines}')
                total_extra_comment_rows += len(wrapped_lines)
                extra_comments.append(wrapped_lines)
            else:
                extra_comments.append([full_comment])
                total_extra_comment_rows += 1

        return extra_comments, total_extra_comment_rows

    def insert_report_comments(self, page_location, current_page , header_size, author_size):
        logger.info(f'Entering insert_report_ending with page_location: {page_location}, current_page: {current_page}')

        # inset the matrix comment
        page_location = self.insert_comment(page_location, self.comment)
        page_location +=2

        # check if can insert footer comments in the remaining page location before inserting signatures
        remaining_page_size = (current_page+1 * self.page_size) - (header_size * current_page) - (author_size)
        extra_comments, total_extra_comment_rows = self.determine_extra_comments()

        logger.debug(f'remaining_page_size: {remaining_page_size}')
        logger.debug(f'self.extra_comment_rows: {self.extra_comment_rows}')
        logger.debug(f'extra_comments: {extra_comments}, total_extra_comment_rows: {total_extra_comment_rows}')

        insert_author_cols = [3,6] if len(self.authors) > 1 else [6]

        # if there is no extra comments
        if(total_extra_comment_rows == 1):
            self.insert_signature(page_location, insert_author_cols, self.authors)
            return

        insert_next_page = False

        # check if need to insert onto next page since no room for extra comments
        if((remaining_page_size - total_extra_comment_rows) < 6):
            self.insert_signature(page_location, insert_author_cols, self.authors)

            insert_next_page = True
            current_page += 1
            next_page_start = (self.page_size * current_page) - (header_size * (current_page-1)) + 1
            page_location = next_page_start

        comment_row = self.ws.cell(row=page_location, column=1)
        comment_row.value = 'Comments:'
        page_location += 1

        for comment in extra_comments:
            for line in comment:
                comment_row = self.ws.cell(row=page_location, column=1)
                comment_row.value = line
                page_location+=1

        if(not insert_next_page):
            page_location += 2
            self.insert_signature(page_location, insert_author_cols, self.authors)


    def insert_tests_info(self, page_location, sample_placement, total_tests, total_hidden_rows):
        logger.info(f'Entering insert_tests_info with sample_placement: {sample_placement}')

        counter = page_location

        # Insert test info (tests, units, recovery)

        for i, test_name in enumerate(self.test_info):
            if(self.hidden_rows[i] == 1):
                continue
            logger.info(f'i: {i}, test_name: {test_name}')

            side_comment = self.side_comments[i] if self.side_comments[i] else ''

            self.set_cell_value_with_format(counter, 1, test_name, Alignment(horizontal='left', vertical='center'), Border(right=self.thin_border_style) )
            self.set_cell_value_with_format(counter, 2, self.units[i])
            self.set_cell_value_with_format(counter, 7, 'ND') # Lab Blank
            self.set_cell_value_with_format(counter, 8, self.so[i] if not is_float(self.so[i]) else significant_figures_convert(float(self.so[i]))) # So Value
            self.set_cell_value_with_format(counter, 9, self.recovery[i])
            self.set_cell_value_with_format(counter, 10, side_comment, Alignment(horizontal='left', vertical='center', indent=1), Border(left=self.thin_border_style))
            counter += 1


        # Insert sample data
        for i, sample in enumerate(sample_placement, start=3):
            print(f'i:{i} sample: {sample} | {sample_placement}')
            current_results = self.sample_data[sample]

            hidden_counter = 0

            # go through all of the rows
            for row in range(total_tests + total_hidden_rows):

                if(self.hidden_rows[row] == 1):
                    hidden_counter += 1
                    continue

                current_sample = self.ws.cell(row=page_location + row - hidden_counter, column=i)
                current_sample.alignment = self.centered_alignment

                # Handle both numeric and string types
                test_val = convert_to_float(current_results[row])
                upper_limit = convert_to_float(self.upper_limits[row])
                lower_limit = convert_to_float(self.lower_limits[row])

                logger.debug(f'test_val: {test_val}, lower_limit: {lower_limit}, upper_limit: {upper_limit}')

                # add the row to the list that will trigger ending comments
                if(row not in self.extra_comment_rows):
                    if(isinstance(upper_limit, float) and isinstance(test_val, float)):
                        if(test_val >= upper_limit):
                            self.extra_comment_rows.append(row)

                    if(isinstance(lower_limit, float) and isinstance(test_val, float)):
                        if(test_val <= lower_limit and row not in self.extra_comment_rows):
                            self.extra_comment_rows.append(row)


                current_sample.value = significant_figures_convert(test_val) if is_float(test_val) else test_val

                current_sample.border = self.thin_side_border

        page_location += total_tests

        # Add borders at the bottom of the table
        for i in range(1, self.total_cols + 1):
            self.ws.cell(row=counter, column=i).border = Border(top=Side(border_style="thin", color="000000"))

        page_location += 1  # Increment the counter for the next row

        return page_location

def convert_to_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return value