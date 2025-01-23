import os

from base_logger import logger

from openpyxl.styles import Font, Alignment, borders, Border, Side

from modules.utils.logic_utils import is_float
from modules.utils.file_utils import get_path_from_json

from pages.reports_page.reports.ExcelReports import ExcelReports

#TODO: set the upper limit comments
#TODO: not scanning the entering in client info

class IcpExcelReport(ExcelReports):

    def __init__(self, param_num, client_info, jobNum, authors, comment, sample_names, sample_data, elements, symbols, limits, units, report_type=1):
        super().__init__(jobNum, report_type)

        self.param_num = param_num
        self.client_info = client_info
        self.authors = authors
        self.comment = comment #TODO: rename this
        self.sample_names = sample_names
        self.sample_data = sample_data
        self.elements = elements
        self.symbols = symbols
        self.limits = limits #is a list
        self.units = units

        self.footer_comments = []

        for key, value in self.sample_data.items():
            print(key, len(value), value)


        print(self.__repr__)

        self.setup()

    def __repr__(self):
        # Limit dictionary output to the first few items for readability
        client_info_preview = {k: v for i, (k, v) in enumerate(self.client_info.items()) if i < 3}
        sample_names_preview = {k: v for i, (k, v) in enumerate(self.sample_names.items()) if i < 3}
        sample_data_preview = {k: v for i, (k, v) in enumerate(self.sample_data.items()) if i < 3}

        return (
            f"IcpExcelReport("
            f"client_info={client_info_preview}, "
            f"jobNum={self.jobNum}, "
            f"authors={self.authors}, "
            f"comments='{self.comment}', "
            f"sample_names={sample_names_preview}, "
            f"sample_data={sample_data_preview}, "
            f"elements={self.elements}, "
            f'symbols={self.symbols}'
            f"limits={self.limits}, "
            f"units={self.units}, "
            f"report_type={self.report_type})"
        )

    # define preferences (water vs soil)
    # show list or show less
    # hardness levels
    # ph comments

    def setup(self):
        logger.info('Entering setup for IcpExcelReport')

        self.total_samples = len(self.sample_data)

        # prepare excel setup page
        self.excel_page_setup()
        # set the excel headers and footers
        self.set_headers_and_footers()
        # format all of the rows to be the correct size
        self.format_rows(self.total_samples)
        # insert client information
        self.insert_client_info(self.client_info, 'D')


    def create_report(self):
        logger.info('Entering create_report')

        sample_sections, sample_placement = self.format_sample_header_names(self.sample_names)

        total_pages = len(sample_placement)
        page_location = 9
        used_samples = 0

        for current_page in range(total_pages):
            sample_amount = len(sample_placement[current_page])

            logger.info(f'sample_amount: {sample_amount}')

            page_location = (61 * current_page) - (8 * (current_page-1)) + 1

            if(current_page == 0):
                page_location = 9

            logger.info(f'page_location: {page_location}')

            page_location = self.insert_sample_name(page_location, sample_sections[current_page])
            page_location = self.insert_test_titles(page_location, sample_amount, used_samples)
            page_location = self.insert_tests_info(page_location, sample_placement[current_page])

            if((current_page+1) == total_pages):
                page_location = self.insert_comment(page_location, self.comment)
                page_location += 1
                self.insert_extra_comments(page_location, current_page )

            else:
                self.insert_next_page_comment(page_location)

            used_samples += sample_amount

        export_path = get_path_from_json('default_paths.json', 'reportsPath')

        file_name = f'W{self.jobNum}_icp.xlsx'
        file_path = os.path.join(export_path, file_name)

        self.save_excel(file_path)

        return file_path, file_name


    def insert_tests_info(self, page_location, sample_placement):
        logger.info('Entering insert_tests_info')

        total_element_rows = len(self.elements)

        # insert elements, symbols, units and max limits comments
        self.populate_element_info(page_location, total_element_rows)

        # insert the sample info
        self.populate_sample_data(page_location, total_element_rows, sample_placement)

        page_location += total_element_rows

        # insert hardness/ph if applicable
        page_location = self.handle_additional_rows(page_location, sample_placement)

        self.add_bottom_border(page_location)

        page_location +=1;

        return page_location


    def populate_element_info(self, current_row, total_rows):

        for index in range(total_rows):
            element_name_row = self.ws.cell(row=current_row, column=1)
            element_symbol_row = self.ws.cell(row=current_row, column=2)
            unit_row = self.ws.cell(row=current_row, column=7)
            limit_comment_row = self.ws.cell(row=current_row, column=8)

            # insert element name and symbol
            element_name_row.value = f'{index+1}) {self.elements[index].capitalize()}'
            element_symbol_row.value = f'{self.symbols[index].capitalize()}'

            unit_row.value = self.units[index]

            # insert the side comment
            upper_limit_str = self.limits[index][2]
            side_comment = self.limits[index][3]

            # Initialize limit comment value as default
            limit_comment_row.value = 'no limit listed'

            if(side_comment != ''):
                limit_comment_row.value = side_comment

            elif(is_float(upper_limit_str)):
                limit_comment_row.value = significant_figures_convert(float(upper_limit_str)) + ' ' + self.units[index]

            #add border and center items
            element_name_row.border = self.thin_right_border
            element_symbol_row.border = self.thin_right_border
            unit_row.border =  self.thin_side_border

            element_symbol_row.alignment = Alignment(horizontal='center', vertical='center')
            unit_row.alignment = Alignment(horizontal='center', vertical='center')
            limit_comment_row.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            current_row +=1

    def populate_sample_data(self, page_location, total_elements, sample_placement):
        logger.info(f'Entering populate_sample_data with page_location: {page_location}, total_elements: {total_elements}, sample_placement: {sample_placement}')

        # insert the sample info
        for current_col, current_sample in enumerate(sample_placement, start=3):
            logger.debug(f'current_col: {current_col}, sample: {current_sample}')

            for current_row in range(0, total_elements):

                sample_cell = self.ws.cell(row=page_location+current_row, column=current_col)
                sample_cell.alignment = Alignment(horizontal='center', vertical='center')
                sample_cell.border = self.thin_right_border

                current_value = self.sample_data[current_sample][current_row]

                lower_limit_str = self.limits[current_row][1]
                upper_limit_str = self.limits[current_row][2]

                # Convert limits to float if they're not empty
                lower_limit = float(lower_limit_str) if lower_limit_str else None
                upper_limit = float(upper_limit_str) if upper_limit_str else None

                if is_float(current_value):
                    current_value_float = float(current_value)

                    if lower_limit is not None:
                        if current_value_float < lower_limit:
                            sample_cell.value = f'< {lower_limit:.3f}'
                        else:
                            # If within limits or no specific action, maintain original conversion
                            sample_cell.value = significant_figures_convert(current_value_float)
                    else:
                        # No valid lower limit; just convert normally
                        sample_cell.value = significant_figures_convert(current_value_float)

                    # trigger to show footer comment
                    if(upper_limit is not None):
                        if(current_value_float > upper_limit):
                            if(current_row not in self.footer_comments):
                                self.footer_comments.append(current_row)
                else:
                    if lower_limit is not None:
                        sample_cell.value = f'< {lower_limit:.3f}'
                    else:
                        sample_cell.value = 'ND'


    def handle_additional_rows(self, page_location, sample_placement):
        logger.info(f'Entering handle_additional_rows with page_location:{page_location}, sample_placement: {sample_placement}')

        total_elements = len(self.elements)
        total_sample_rows = len(self.sample_data[sample_placement[0]])
        additional_rows = total_sample_rows - total_elements
        #additional_rows = 1 if self.param_num not in [17, 14, 6] else 2

        logger.debug(f'additional_rows: {additional_rows}')

        # hardness only for liquids not solids
        if(additional_rows == 2):
            hardness_values = []

            self.insert_hardness_cols(page_location)

            for col, sample in enumerate(sample_placement, start=3):
                hardness_value = self.sample_data[sample][-2]
                hardness_value = float(hardness_value) if hardness_value.isnumeric() else 0
                hardness_values.append(hardness_value)
                self.insert_value_to_cell(page_location, col, hardness_value)

            logger.debug(f'hardness_values: {hardness_values}')

            # insert hardness side comment
            cell = self.ws.cell(row= page_location, column=8, value=hardness_levels_comment(max(hardness_values)))
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ph_values = []
        ph_row = page_location + additional_rows - 1

        self.insert_ph_cols(ph_row)

        for col, sample in enumerate(sample_placement, start=3):
            ph_value = self.sample_data[sample][-1]
            ph_values.append(ph_value)
            self.insert_value_to_cell(ph_row, col, ph_value)

        logger.debug(f'ph_values: {ph_values}')

        # insert pH side comment
        cell = self.ws.cell(row= ph_row, column=8, value='7.0 to 10.5')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)


        page_location += additional_rows
        return page_location

    def insert_value_to_cell(self, row,  column, value):
        cell = self.ws.cell(row=row, column=column)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_right_border

        try:
            float_version = float(value)
            cell.value = significant_figures_convert(float_version)

        except Exception:
            logger.warning(f'Invalid value: {value}')
            cell.value = 'ND'

    def insert_hardness_cols(self, page_location):
        self.insert_parameter_detail(page_location, 'Hardness', 'CaCO₃', 'mg/L', '')

    def insert_ph_cols(self, page_location):
        self.insert_parameter_detail(page_location, 'pH', '', 'units', '7.0 to 10.5')

    def insert_parameter_detail(self, page_location, element_name, symbol, unit, comment):
        cells = {
            'element_cell': self.ws.cell(row=page_location, column=1, value=element_name),
            'symbol_cell': self.ws.cell(row=page_location, column=2, value=symbol),
            'unit_cell': self.ws.cell(row=page_location, column=7, value=unit),
            'comment_cell': self.ws.cell(row=page_location, column=8, value=comment),
        }

        # Set alignments
        for cell_key in ['symbol_cell', 'unit_cell']:
            cells[cell_key].alignment = Alignment(horizontal='center', vertical='center')
        cells['comment_cell'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # Set borders
        cells['element_cell'].border = self.thin_right_border
        cells['symbol_cell'].border = self.thin_right_border
        cells['unit_cell'].border = self.thin_side_border


    def determine_extra_comments(self):
        logger.info(f'Entering determine_extra_comments')
        logger.info(f'footer_comments: {self.footer_comments}')

        total_extra_comment_rows = 1
        extra_comment_lines = []

        for row in self.footer_comments:
            total_row_size = 2
            element_name = self.limits[row][0]
            comment = self.limits[row][5]

            full_comment = f'{element_name.upper()}: {comment}'

            total_row_size += len(full_comment)

            logger.info(f'total_row_size: {total_row_size}')

            if(total_row_size > self.row_char_limit):
                wrapped_lines = split_sentence_by_words(full_comment)
                logger.info(f'wrapped_lines:{wrapped_lines}')
                total_extra_comment_rows += len(wrapped_lines)
                extra_comment_lines.append(wrapped_lines)

            else:
                extra_comment_lines.append([full_comment])
                total_extra_comment_rows += 1

        return extra_comment_lines, total_extra_comment_rows

    #TODO: need to factor in for total pages with the comments sizes
    def insert_extra_comments(self, page_location, current_page):
        logger.info(f"Entering insert_extra_comments page_location: {page_location}, current_page: {current_page}")

        extra_comments, comments_rows_length = self.determine_extra_comments()

        #TODO: check what the pH scores comment
        if(self.param_num == 17):
            pass

        logger.info(f'comment_rows: {comments_rows_length}')

        temp = page_location + 6 #comment, line, line, line, name, position

        insert_author_cols = [3,6] if len(self.authors) > 1 else [6]

        # check if there is extra comments to insert
        if(comments_rows_length == 1):
            #FIXME: if the correct parameter is selected
            comment_row = self.ws.cell(row=page_location, column=1)
            comment_row.value = 'All constituents tested meet Canadian and B.C. drinking water standards.'
            page_location += 3

            # insert signature
            self.insert_signature(page_location, insert_author_cols, self.authors)
            return

        #TODO: check if there is room to insert the comment here

        current_page_length = (current_page +1) * self.page_size
        expected_length = temp + comments_rows_length

        logger.info(f'expected_length: {expected_length}, current_page_length: {current_page_length}')

        insert_next_page = False #TODO: rename this variable

        if(expected_length >= current_page_length):

            page_location += 1

            self.insert_signature(page_location, insert_author_cols, self.authors)

            #insert comments on next page
            page_location = (self.page_size * (current_page + 1) + 1 )
            total_rows = (self.page_size * (current_page + 2))

            logger.info(f'new start: {page_location}')

            insert_next_page = True

        comment_row = self.ws.cell(row=page_location, column=1)
        comment_row.value = 'Comments:'
        page_location += 1;

        for comment in extra_comments:
            for line in comment:
                comment_row = self.ws.cell(row=page_location, column=1)
                comment_row.value = line
                page_location+=1

        if(not insert_next_page):
            page_location += 2

            self.insert_signature(page_location, insert_author_cols, self.authors)

def split_sentence_by_words(sentence, max_chars_per_line=140):
    words = sentence.split()
    lines = []
    current_line = ""

    for word in words:

        if len(current_line) + len(word) + 1 <= max_chars_per_line:  # +1 for the space
            current_line += word + " "
        else:
            lines.append(current_line.strip())  # Add the current line to the list
            current_line = word + " "  # Start a new line with the current word

    if current_line:  # Add the last line if it's not empty
        lines.append(current_line.strip())

    return lines

def hardness_levels_comment(max_val):
    if(max_val <= 75):
        return '0-75 mg/L = soft'
    if(max_val <= 150):
        return '75-150 mg/L = medium'
    if(max_val <= 300):
        return '150-300 mg/L = hard'
    else:
        return '> 300 mg/L = very hard'

def significant_figures_convert(value):
    if(value >= 100):
        return (f'{value:.0f}')
    if(value >=10):
        return (f'{value:.1f}')
    if(value >= 1):
        return (f'{value:.2f}')
    if(value < 1):
        return (f'{value:.3f}')

    return value