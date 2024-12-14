import os
import math
import platform

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintOptions

from base_logger import logger


from modules.constants import elementSymbols
from modules.utils.pickle_utils import load_pickle

class ExcelReports:

    def __init__(self, jobNum, report_type):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.jobNum = jobNum

        if platform.system() == "Windows":
            self.system = 1
        elif platform.system() == "Darwin":  # MacOS
            self.system = 2

        print(f'self.system : {self.system}')

        #1 = ICP, 2 = CHM
        self.report_type = report_type

        self.total_cols = 8
        self.page_size = 61

        self.default_styles()

    def default_styles(self):
        self.default_font = Font(name="Times New Roman", size=9)
        self.thin_border_style = Side(border_style="thin", color="000000")
        self.double_border_style = Side(border_style='double', color="000000")

        self.thin_border = Border(right=self.thin_border_style, left=self.thin_border_style)
        self.left_right_border = Border(right=self.thin_border_style, left=self.thin_border_style)

        self.centered_alignment = Alignment(horizontal='center', vertical='center')

    def excel_page_setup(self):
        logger.info('Entering excel_page_setup')

        self.ws.sheet_view.view = "pageLayout"

        # set the excel page settings
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0  # Automatic height fitting

        # Set to landscape or portrait as needed
        self.ws.page_setup.orientation = self.ws.ORIENTATION_PORTRAIT

        #Set paper size
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4  # Set paper size to A4 or applicable size

        # Ensure that the scaling happens automatically
        self.ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=True)

        # set custom page margins
        self.ws.page_margins = PageMargins(
            left=0.7,
            right=0.7,
            top=0.75,
            bottom=0.75,
            header=0.3,
            footer=0.3
        )

        # set column dimensions
        self.ws.column_dimensions['A'].width = 20 #120px
        self.ws.column_dimensions['H'].width = 19
        self.ws.print_title_rows = '1:8' # the first two row

    def format_rows(self, total_samples):
        logger.info(f'Entering format_rows with total_samples: {total_samples}')

        # Mac System
        if(self.system == 2):
            window_conversion = 1
            row_height_pixels = 13
        else:
            window_conversion = 0.75
            row_height_pixels = 15

        total_pages = math.ceil(total_samples/4)
        total_rows = (self.page_size * total_pages) - (8 * (total_pages -1))

        logger.debug(f'total_pages: {total_pages}, total_rows: {total_rows}')

        # iterate through all of the rows and set them the correct height
        for row in self.ws.iter_rows(min_row=1, max_col=self.total_cols, max_row=total_rows):
            for cell in row:
                cell.font = self.default_font
                self.ws.row_dimensions[cell.row].height = (row_height_pixels * window_conversion)

    def set_headers_and_footers(self):
        logger.info(f'Entering create_footer')

        if(self.report_type == 1):
            title = 'ICP Report'
        else:
            title = 'CHM Report'

        left_text = f"{title}: &D"
        right_text_header = f"Page &P of &N \n W{self.jobNum}"
        right_text_footer = '&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
        center_text_footer = "&B MB Laboratories Ltd.&B \nwww.mblabs.com"
        contact_footer = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com'

        # insert the headers
        self.insert_header_footers(self.ws.oddHeader, left_text, right_text_header)
        self.insert_header_footers(self.ws.evenHeader, left_text, right_text_header)

        # insert the footers

        self.ws.oddFooter.font_name = 'Times New Roman'
        self.ws.oddFooter.font_size = 11
        self.ws.evenFooter.font_name = 'Times New Roman'
        self.ws.evenFooter.font_size = 11

        self.insert_header_footers(self.ws.oddFooter, contact_footer, right_text_footer, center_text_footer)
        self.insert_header_footers(self.ws.evenFooter, contact_footer, right_text_footer, center_text_footer)


    def insert_header_footers(self, placement, left_text, right_text, center_text=None):

        font = 'Times New Roman'
        font_size = 11

        placement.left.font = font
        placement.left.size = font_size
        placement.left.text = left_text

        placement.right.font = font
        placement.right.size = font_size
        placement.right.text = right_text

        if center_text:
            placement.center.font = font
            placement.center.size = font_size
            placement.center.text = center_text

    def insert_client_info(self, client_info, column2):
        logger.info('Entering insert_client_info')

        if not self.ws:
            raise ValueError("Worksheet is not set. Use set_worksheet() to initialize the worksheet.")

        # Primary client details
        self.ws['A1'] = client_info.get('clientName', '')

        self.ws['A2'] = client_info.get('attn', '*') or '*'
        self.ws['A3'] = client_info.get('addy1', '')
        self.ws['A4'] = f"{client_info.get('addy2', '')}, {client_info.get('addy3', '')}"

        self.ws['A6'] = f"TEL: {client_info.get('tel', '')}"
        self.ws['A7'] = client_info.get('email', '')

        # Secondary client details
        self.ws[f'{column2}1'] = f"Date: {client_info.get('date', '')}  ({client_info.get('time', '')})"
        self.ws[f'{column2}2'] = f"Source: {client_info.get('sampleType1', '')}"
        self.ws[f'{column2}3'] = f"Type: {client_info.get('sampleType2', '')}"
        self.ws[f'{column2}4'] = f"No. of Samples: {client_info.get('totalSamples', '')}"
        self.ws[f'{column2}6'] = f"Arrival temp: {client_info.get('recvTemp', '')}"
        self.ws[f'{column2}7'] = f"PD: {client_info.get('payment', '')}"

        return self.ws

    def format_sample_header_names(self, sample_names):
        ''' combines the sample_name with it's time and date values into one string'''
        logger.info('Entering format_sample_header_names')

        sample_sections = []
        sample_placement = []

        display_limit = 4

        current_word = ''
        temp = []

        for i, (key,value) in enumerate(sample_names.items(), start=1):
            stripped_word = " ".join(value.split())
            current_word += " " + str(i) + ") " + stripped_word + " "
            temp.append(key)

            if(i % display_limit == 0):
                sample_sections.append(current_word)
                sample_placement.append(temp)
                current_word = ""
                temp = []

        if(current_word != ''):
            sample_sections.append(current_word)
            sample_placement.append(temp)

        logger.debug(f'sample_sections: {sample_sections}, sample_placement:{sample_placement}')

        return sample_sections, sample_placement

    def insert_sample_name(self,row, sample_section):
        logger.info(f'Entering insert_sample_name with row: {row}, sample_section: {sample_section}')

        sample_cell = self.ws.cell(row=row, column=1)
        sample_cell.value = f'Samples: {sample_section}'
        sample_cell.border = Border(bottom=self.thin_border_style)

        self.ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=self.total_cols)
        sample_cell.alignment = Alignment(wrap_text=True)

        return row + 3

    def set_cell_value_with_alignment(self, row, col, value, alignment=None):
        logger.info('Entering set_cell_value_with_alignment')
        if(alignment is None):
            alignment = self.centered_alignment

        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = alignment
        return cell

    def set_cell_value_with_format(self, row, col, value, alignment=None, border=None):
        logger.info('Entering set_cell_value_with_format')

        if(alignment is None):
            alignment = self.centered_alignment

        if(border is None):
            border = self.thin_border

        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = alignment
        cell.border = border
        return cell

    def set_sample_names(self, allowed_borders: list, page_location: int, total_samples: int, start_val: int):

        for i in range(total_samples):
            sample = self.ws.cell(row=page_location, column=i + 3)
            allowed_borders.append(i+3)
            sample.value = f"Sample {start_val + i + 1}" if start_val != 0 else f"Sample {i + 1}"
            sample.alignment = self.centered_alignment
            sample.border = self.thin_border

        return page_location

    def insert_test_titles(self, page_location: int, total_samples: int, start_val: int):
        logger.info(f'Entering insert_test_titles with page_location: {page_location}, total_samples:{total_samples}, start_val: {start_val}')

        titles = {
            0: {'col_1': 'Tests Names', 'col_2': 'Units', 'col_7':'Recovery', 'col_8': 'Comment'},
            1: {'col_1': 'Elements', 'col_2': 'Symbols', 'col_7': 'Units', 'col_8': 'Maximum Limits'},
        }

        # Insert test type in column 1
        self.set_cell_value_with_alignment(page_location, 1, titles[self.report_type]['col_1'] , Alignment(horizontal='left', vertical='center'))

        # Insert units/symbols in column 2
        unit_cell = self.set_cell_value_with_alignment(page_location, 2, titles[self.report_type]['col_2'])
        unit_cell.border = self.thin_border

        # Apply sample names and update page_location
        allowed_borders = [1, 2, 7, 8]
        self.set_sample_names(allowed_borders, page_location, total_samples, start_val)

        # Insert comment/limits in column 7
        recv_cell = self.set_cell_value_with_alignment(page_location, 7, titles[self.report_type]['col_7'])
        recv_cell.border = self.thin_border

        # Insert comment or maximum limits in column 8
        self.set_cell_value_with_alignment(page_location, 8, titles[self.report_type]['col_8'], Alignment(horizontal='left', vertical='center', indent=1))

        page_location +=1

        return self.apply_borders_and_formatting(allowed_borders, page_location)

    def apply_borders_and_formatting(self, allowed_borders, page_location):
        """
        Helper function to apply borders and formatting to rows based on allowed borders.
        """

        for i in range(1, self.total_cols + 1):
            current = self.ws.cell(row=page_location, column=i)
            if i in allowed_borders:
                if i != 8:
                    current.border = Border(right=self.thin_border_style, bottom=self.double_border_style)
                else:
                    current.border = Border(bottom=self.double_border_style)
                if i == 7:
                    current.border = Border(right=self.thin_border_style, left=self.thin_border_style, bottom=self.double_border_style)
            else:
                current.border = Border(bottom=self.double_border_style)

        page_location +=1
        return page_location

    def insert_comment(self, page_location, comment):
        logger.info('Entering insert_comment')

        for _, commentLine in enumerate(comment):
            temp = self.ws.cell(row = page_location, column=1)
            temp.value = commentLine
            page_location+=1

        return page_location

    def insert_next_page_comment(self, page_location):
        logger.info('Entering insert_next_page_comment')
        comment =self. ws.cell(row=page_location, column=1)
        comment.value = 'continued on next page....'
        comment.font = Font(bold=True, size=9, name="Times New Roman")


    def insert_signature(self,  page_location:int , start_col: int, author_info: list) -> int:
        logger.info(f'Entering insert_signature with page_location:{repr(page_location)}, start_col: {repr(start_col)}, author_info: {repr(author_info)}')

        authorNames = []
        authorRoles = []

        for author in author_info:
            authorNames.append(author[1])
            authorRoles.append(author[2])

        for i, col in enumerate(start_col):
            scientistNamePos = self.ws.cell(row=page_location, column=col)
            scientistRolePos = self.ws.cell(row=page_location+1, column=col)

            scientistNamePos.value = authorNames[i]
            scientistRolePos.value = authorRoles[i]

            for j in range(2):
                signatureLine = self.ws.cell(row=page_location, column=col+j)
                signatureLine.border = Border(top=self.thin_border_style)

    def save_excel(self, filePath):

        self.wb.save(filePath)

class ChmExcelReport(ExcelReports):

    def __init__(self, client_info, jobNum, authors, comment, sample_names, sample_data, test_info, units, recovery, report_type=1):
        super().__init__(jobNum, report_type)
        self.authors = authors
        self.client_info = client_info
        self.comment = comment
        self.sample_names = sample_names
        self.sample_data = sample_data
        self.test_info = test_info
        self.units = units
        self.recovery = recovery

        print(f'sample_data: {sample_data}')

        self.setup()

    def setup(self):
        logger.info('Entering setup for ChmExcelReport')

        self.total_samples = len(self.sample_data)

        self.excel_page_setup()
        self.set_headers_and_footers()
        self.format_rows(self.total_samples)

        #note can insert
        self.insert_client_info(self.client_info, 'D')

    def create_report(self):
        logger.info('Entering create_report')

        total_tests = len(self.test_info)

        allocated_space = 20;
        table_size = 6 + total_tests

        sample_sections, sample_placement = self.format_sample_header_names(self.sample_names)

        print(f'sample_placement: {sample_placement}')

        total_sample_sections = len(sample_sections)
        total_tables_with_comments = math.floor((self.page_size - allocated_space)/table_size)
        total_pages = math.ceil(total_sample_sections/total_tables_with_comments)

        #page_location = 9;
        used_samples = 0
        current_sample = 0

        logger.info('Preparing to write excel file')
        for current_page in range(total_pages):
            logger.debug(f'current_page: {current_page} / {total_pages}')

            sample_amount = len(sample_placement[current_sample]) # how many samples in the current batch

            if(current_page != 0):
                # onto new page but set the line to start where on equivalent of 9 on new page
                page_location = (self.page_size * current_page) - (8 * (current_page -1)) +1
            else:
                page_location = 9

            logger.debug(f'page_location: {page_location}')

            if(current_page+1 == total_pages):
                logger.debug('Last Page')

                remaining_samples = total_sample_sections - current_sample

                for i in range(remaining_samples):
                    sample_amount = len(sample_placement[current_sample])

                    page_location = self.insert_sample_name(page_location, sample_sections[current_sample])
                    page_location = self.insert_test_titles(page_location, sample_amount, used_samples)
                    page_location = self.insert_tests_info(page_location, sample_placement[current_sample], total_tests)

                    if(i+1 == remaining_samples):
                        page_location = self.insert_comment(page_location, self.comment)
                        page_location +=2

                        if(len(self.authors) > 1):
                            self.insert_signature(page_location, [3,6], self.authors)
                        else:
                            self.insert_signature(page_location,[6], self.authors)

                    current_sample +=1
                    used_samples += 4

            else:
                logger.debug('Not Last Page')
                page_location = self.insert_sample_name(page_location, sample_sections[current_sample])
                page_location = self.insert_test_titles(page_location, sample_amount, used_samples)
                page_location = self.insert_tests_info(page_location, sample_placement[current_sample], total_tests)


                if(i+1 == total_tables_with_comments):
                    self.insert_next_page_comment(page_location)

                current_sample +=1
                used_samples += 4


            file_path_manager = load_pickle('data.pickle')
            export_path = file_path_manager['reportsPath']

            file_name = 'W' + str(self.jobNum) + ".chm"
            file_path = os.path.join(export_path, file_name)

            self.save_excel(file_path)

            return file_path, file_name


    def insert_tests_info(self, page_location, sample_placement, total_tests):
        logger.info(f'Entering insert_tests_info with sample_placement: {sample_placement}')

        counter = page_location

        # Insert test info (tests, units, recovery)
        for i in range(len(self.test_info)):
            test_value = self.test_info[i] if self.test_info[i] else 'Error'
            self.set_cell_value_with_format(counter, 1, test_value, Alignment(horizontal='left', vertical='center'))
            self.set_cell_value_with_format(counter, 2, self.units[i])
            self.set_cell_value_with_format(counter, 7, self.recovery[i])
            self.set_cell_value_with_format(counter, 8, '', Alignment(horizontal='left', vertical='center', indent=1))
            counter += 1

        # Insert sample data
        for i, sample in enumerate(sample_placement, start=3):
            print(f'i:{i} sample: {sample} | {sample_placement}')
            current_results = self.sample_data[sample]

            # go through all of the rows
            for j in range(total_tests):
                current_sample = self.ws.cell(row=page_location + j, column=i)
                current_sample.alignment = self.centered_alignment

                # Handle both numeric and string types
                value = float(current_results[j]) if isinstance(current_results[j], (int, float)) else current_results[j]

                current_sample.value = value
                current_sample.border = self.thin_border

        page_location += total_tests

        # Add borders at the bottom of the table
        for i in range(1, 9):
            self.ws.cell(row=counter, column=i).border = Border(top=Side(border_style="thin", color="000000"))

        page_location += 1  # Increment the counter for the next row

        return page_location



class IcpExcelReport(ExcelReports):

    def __init__(self, client_info, jobNum, authors, comments, sample_names, sample_data, elements, limits, units, report_type=1):
        super().__init__(jobNum, report_type)

        self.client_info = client_info
        self.authors = authors
        self.comments = comments
        self.sample_names = sample_names
        self.sample_data = sample_data
        self.elements = elements
        self.limits = limits
        self.units = units

        self.setup()

    def setup(self):
        logger.info('Entering setup for IcpExcelReport')

        self.total_samples = len(self.sample_data)

        # prepare excel setup page
        self.excel_page_setup()
        # set the excel headers and footers
        self.set_headers_and_footers()
        # format all of the rows to be the correct size
        self.format_rows(self.total_samples)
        # insert client information
        self.insert_client_info(self.client_info, 'D')

    def create_report(self):
        logger.info('Entering create_report')

        sample_sections, sample_placement = self.format_sample_header_names(self.sample_names)

        total_pages = len(sample_placement)
        page_location = 9
        used_samples = 0

        for current_page in range(total_pages):
            sample_amount = len(sample_placement[current_page])

            logger.info(f'sample_amount: {sample_amount}')

            # first page information
            if(current_page == 0):
                page_location = 9
            else:
                pass

    def insert_icp_document_end(ws, page_location, footer_comment, author_info):
        pass


    def insert_tests_info(self):
        pass

    def significant_figures_convert(self, value):
        if(value >= 100):
            return (f'{value:.0f}')
        if(value >=10):
            return (f'{value:.1f}')
        if(value >= 1):
            return (f'{value:.2f}')
        if(value < 1):
            return (f'{value:.3f}')

        return value;