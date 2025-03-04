import csv
import os
import re
import json
import openpyxl
import string

from base_logger import logger
from copy import copy
from datetime import date

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QFileDialog, QPushButton, QTableWidgetItem, QTableWidget , QVBoxLayout, QDialog,  QSizePolicy, QSizeGrip )

from modules.dialogs.basic_dialogs import error_dialog
from modules.utils.file_utils import get_path_from_json

#TODO: edit how the dialogs look for this

def icp_upload(file_path, db):
    logger.info(f'Entering icp_upload with file_path: {file_path}')

    upload_methods = {
        '.txt': process_txt_file,
        '.xlsx': process_xlsx_file,
        '.csv': process_csv_file,
    }

    # Get the file extension using os.path.splitext
    file_extension = os.path.splitext(file_path)[1].lower()

    upload_method = upload_methods.get(file_extension)

    if upload_method:
        try:
            upload_method(db, file_path)
        except Exception as e:
            logger.error(f'Error uploading file: {e}')
            error_dialog('Error Uploading File', f'There was an error uploading file: {file_path}')

    else:
        logger.warning(f'Invalid file type: {file_path}')
        error_dialog('Error Uploading File', f'Invalid file type: {file_extension}')

def process_csv_file(db, file_path):
    logger.info(f'Entering process_csv_file with file_path:{file_path}')

    base_name = os.path.basename(file_path)
    file_name, _ = os.path.splitext(base_name)
    new_name = f'{file_name}.csv'

    output_path = get_path_from_json('default_paths.json', 'ispDataUploadPath')
    new_path = os.path.join(copy(output_path), new_name)


    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)

        for row_index, row in enumerate(reader):
            if(row_index == 0):
                row_length = len(row)

                if(row_length == 1):
                    return process_machine_csv_file(db, file_path, base_name)

                if(row_length > 2):
                    return process_edited_csv_file(db, file_path, base_name)

def process_machine_csv_file(db, file_path, base_name):
    logger.info(f'Entering process_machine_csv_file with file_path: {file_path}')

    symbols = []

    job_numbers, sample_numbers = [], []

    job_data = {}

    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)

        element_data = {}

        for row_index, row in enumerate(reader):

            if(row_index == 1):
                symbols = get_element_symbols(row[6:])

            if(row_index >= 2):
                label_name = row[0]

                if(re.search(r'\d{6}-\d{1,2}$', label_name)):
                    for col, value in enumerate(row[6:]):

                        element_symbol = symbols[col]
                        job_number = label_name.split('-')[0]

                        if(job_number not in job_numbers):
                            job_numbers.append(job_number)

                        sample_numbers.append([label_name, row_index, element_symbol, value])
                        element_data[element_symbol] = value

                    job_data[label_name] = element_data


    for key, value in job_data.items():
        logger.debug(f'{key}, {value}')

    save_status = viewIcpTable(file_path, sample_numbers, reportType=1)
    logger.debug(f"save_status: {repr(save_status)}")

    if(save_status):
        for (key, value) in job_data.items():
            jobNum = key.split('-')[0]
            today_date = date.today()
            package_data = json.dumps(value)

            if(key):
                sql = 'INSERT OR REPLACE INTO icp_upload values(?, ?, ?, ?, ?, 1)'
                db.execute(sql, (key, jobNum, base_name, package_data, today_date))
                db.commit()

        return job_numbers, job_data

def get_element_symbols(symbols):

    element_symbols = []

    for symbol in symbols:
        symbol_split = symbol.split()

        element_symbols.append(symbol_split[0])

    logger.debug(f'{element_symbols}')

    return element_symbols

def process_edited_csv_file(db, file_path, base_name):
    logger.info(f'Entering process_edited_csv_file with file_path: {file_path}')

    base_name = os.path.basename(file_path)
    file_name, _ = os.path.splitext(base_name)
    new_name = f'{file_name}.csv'

    output_path = get_path_from_json('default_paths.json', 'ispDataUploadPath')
    new_path = os.path.join(copy(output_path), new_name)

    job_numbers, sample_numbers = [], []
    job_data, element_data = {}, {}

    current_job = ''

    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)

        for row_index, row in enumerate(reader):

            logger.info(f'row: {row_index}, {row}')

            if(row_index >= 4):
                label_name = row[0]
                date_time = row[2]
                element = row[4]
                unadj_data = row[6]
                concent_data = row[7]
                unit = row[8]

                # make sure that this is a sample
                # FIXME: had some files that appear liked '130088-3 fe etc' in the csv column, fix that seems strange
                if(re.search('\d{6}-\d{1,2}$', label_name)):

                    sample_numbers.append([label_name, row_index, element, unadj_data])


                    if(current_job == ''):
                       current_job = label_name

                    elif(current_job != label_name):
                        job_data[current_job] = element_data
                        element_data = {}
                        current_job = label_name

                    element_data[element] = unadj_data
                    job_number = label_name.split('-')[0]

                    if(job_number not in job_numbers):
                        job_numbers.append(job_number)


        job_data[current_job] = element_data

    for key, value in job_data.items():
        logger.debug(key, value )

    save_status = viewIcpTable(file_path, sample_numbers, reportType=1)
    logger.info(f"save_status: {repr(save_status)}")

    if(save_status):
        for (key, value) in job_data.items():
            jobNum = key.split('-')[0]
            today_date = date.today()
            package_data = json.dumps(value)

            if(key):
                sql = 'INSERT OR REPLACE INTO icp_upload values(?, ?, ?, ?, ?, 1)'
                db.execute(sql, (key, jobNum, base_name, package_data, today_date))
                db.commit()

        return job_numbers, job_data

    return False


#TODO: sort by name
def process_txt_file(db, file_path):
    logger.info(f'Entering process_txt_file with file_path: {file_path}')

    base_name = os.path.basename(file_path)
    file_name, _ = os.path.splitext(base_name)
    new_name = f'{file_name}.csv'

    starting_line = 'Date Time Label Element Label (nm) Conc %RSD Unadjusted Conc Intensity %RSD'
    headers = ['Sample', 'Analyze', 'Element', 'HT', ' ', 'units', 'rep', 'Date', 'Time']

    output_path = get_path_from_json('default_paths.json', 'ispDataUploadPath')
    new_path = os.path.join(copy(output_path), new_name)

    writer = csv.writer(open(new_path, 'w'))
    writer.writerow(headers)

    job_numbers, sample_numbers = [], []
    job_data, element_data = {}, {}

    current_job = ''

    with open(file_path, 'r') as file1:
        lines = file1.readlines()

        processing_data = False

        for i, line in enumerate(lines):
            if(processing_data):

                if(re.search('([1-9]|[1-9][0-9]) of ([1-9]|[1-9][0-9])$', line)):
                    processing_data = False
                    continue

                split_line = line.split()

                if(re.search('\d{6}-\d{1,2}', split_line[2])):
                    sample_date = split_line[0]
                    sample_time = split_line[1]
                    sample_name = split_line[2]
                    element = split_line[3]
                    value = split_line[6]

                    sample_numbers.append([sample_name, i, element, value])

                    if(len(split_line) > 14):
                        unit = split_line[8].strip('()')
                    else:
                        unit = split_line[7].strip('()')

                    row = [sample_name, 1, element, 1, value, unit, 1, sample_date, sample_time]

                    #logger.info(split_line)
                    #logger.info(f'{row}, {len(split_line)}')

                    if(current_job == ''):
                       current_job = sample_name

                    elif(current_job != sample_name):
                        job_data[current_job] = element_data
                        element_data = {}
                        current_job = sample_name

                    element_data[element] = value
                    job_number = sample_name.split('-')[0]

                    if(job_number not in job_numbers):
                        job_numbers.append(job_number)

                    writer.writerow(row)

            # next line where the information begin
            if(line.strip() == starting_line):
                processing_data = True

    job_data[current_job] = element_data

    logger.debug(f'sample_numbers: {sample_numbers}')
    logger.debug(f'job_data: {job_data}')

    save_status = viewIcpTable(file_path, sample_numbers, reportType=1)
    logger.info(f"save_status: {repr(save_status)}")

    if(save_status):
        for (key, value) in job_data.items():
            jobNum = key.split('-')[0]
            today_date = date.today()
            package_data = json.dumps(value)

            if(key):
                sql = 'INSERT OR REPLACE INTO icp_upload values(?, ?, ?, ?, ?, 1)'
                db.execute(sql, (key,jobNum, base_name, package_data, today_date))
                db.commit()

        return job_numbers, job_data

    return False

#scans thought all the text files and finds all the different sample types and the ISP and CHM files
#TODO: insert try catch block
def process_xlsx_file(db, filePath):
    logger.info('Entering process_xlsx_file with parameters: filePath: {filePath}')

    wb = openpyxl.load_workbook(filePath)
    sheets = wb.sheetnames

    baseName = os.path.basename(filePath)
    fname = baseName.split('.xlsx')[0]

    logger.debug(f'FileName: ', fname)

    newName = fname + '_formatted'  + '.xlsx'

    file_path = get_path_from_json('default_paths.json', 'ispDataUploadPath')
    newPath = os.path.join(copy(file_path), newName)

    ws = wb[sheets[0]]

    sampleTypeColumn = ws['E']
    sampleNameColumn = ws['G']

    elementConversion = ['As', 'Se', 'Cd', 'Sb', 'Hg', 'Pb', 'U']
    elementColumns = ['I', 'J', 'M', "Q", 'U', 'AA', 'AC']
    selectedRows = []

    sampleInfo = {}

    for cell in sampleTypeColumn:
        if(cell.value == 'Sample'):

            currentSampleName = ws.cell(row=cell.row, column=7).value
            pattern = r'^\d{6}-\d{3}'

            if(re.search(pattern, currentSampleName)):
                selectedRows.append(cell.row)

                sampleName = formatJobSampleString(currentSampleName)
                jobNum = sampleName[:6]
                sampleData = {}

                for i, element in enumerate(elementColumns):

                    col_index = openpyxl.utils.column_index_from_string(element)
                    elementVal = ws.cell(row=cell.row, column=col_index).value

                    if(elementVal == '<0.000'):
                        sampleData[elementConversion[i]] = 0.00
                    else:
                        sampleData[elementConversion[i]] = elementVal

                sampleInfo[sampleName] = sampleData

    for key, value in sampleInfo.items():
        query  = 'INSERT OR REPLACE INTO icp_upload values(?, ?, ?, ?, ?, 2)'
        jobNum = key[:6]
        todayDate = date.today()
        tempData = json.dumps(value)
        if(jobNum):
            db.execute(query, (key, jobNum, baseName, tempData, todayDate))
            db.commit()

    # Format the machine data
    formatMachineData(ws, selectedRows, elementColumns, newPath)


    #TODO: create a view label for when we upload this
    return;

def formatMachineData(ws, selectedRows, elementColumns, newPath):
    logger.info('Entering formatMachineData with parameters:')
    logger.info(f'selectedRows: {repr(selectedRows)}')
    logger.info(f'elementColumns: {repr(elementColumns)}')
    logger.info(f'newPath: {repr(newPath)}')

    newWb = openpyxl.Workbook()
    ws2 = newWb.active

    ws2.cell(row=1, column=1).value = 'Sample'
    tableNames = ['', 'Rjct', 'Data File', 'Acq. Date-Time', 'Type', 'Level', 'Sample Name', 'Total Dil',
                  '75  As  [ He ] ', '78  Se  [ H2 ] ', '111  Cd  [ No Gas ] ', '123 Sb [He]', '202 Hg [He]', '208 Pb [He]', '238 U[He]'
                  ]

    ws2.merge_cells('A1:H1')

    column_num = 1;
    for item in tableNames:
        ws2.cell(row=2, column=column_num).value = item;
        column_num += 1;

    row_num = 3;
    for row_value in selectedRows:

        for currentPos in range(1,8):
            ws2.cell(row=row_num, column=currentPos).value = ws.cell(row=row_value, column=currentPos).value

        tempCol = 9
        for col in elementColumns:
            col_index = openpyxl.utils.column_index_from_string(col)
            ws2.cell(row=row_num, column=tempCol).value = ws.cell(row=row_value, column=col_index).value
            tempCol+=1;


        row_num+=1;

    newWb.save(newPath)

def formatJobSampleString(inputString):
    sample = inputString.strip()
    match = re.match(r'(\d+)-0+(\d+)', sample)

    if match:
        first_part = match.group(1)
        second_part = match.group(2)
        formatted_string = f"{first_part}-{second_part}"

        return formatted_string;

def formatStringArray(inputArray):
    outputArray = []

    for sample in inputArray:
        sample = sample.strip()

        # Use regular expressions to extract the desired pattern
        match = re.match(r'(\d+)-0+(\d+)', string)

        if match:
            # Get the captured groups from the regex match
            first_part = match.group(1)
            second_part = match.group(2)

            # Construct the desired format
            formatted_string = f"{first_part}-{second_part}"
            outputArray.append(formatted_string)

    return(outputArray)

def viewIcpTable(filePath, data, reportType):
    logger.info(f'Entering viewIcpTable with parameters: filePath: {filePath}, data: {data}, reportType: {reportType}')

    dialog = icpTableView(filePath, data, reportType)

    if dialog.exec_() == QDialog.Accepted:
        return True;
    else:
        return False;

#******************************************************************
#    Classes
#******************************************************************


class icpTableView(QDialog):
    def __init__(self, filePath, data, reportType):
        super().__init__()

        self.resize(600, 500)
        self.setWindowTitle(filePath)

        headers = ['Sample Number', 'Line', 'Element', 'Value', 'duplicate Line']

        layout = QVBoxLayout()

        table_widget = QTableWidget()
        table_widget.setRowCount(len(data))
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)

        seen = {}

        for row, rowData in enumerate(data):
            sampleNum = rowData[0]
            line = rowData[1]
            parameterName = rowData[2]

            if (sampleNum, parameterName) in seen:
                previousRow = seen[(sampleNum, parameterName)]
                item = QTableWidgetItem(str(previousRow))
                table_widget.setItem(row,4, item)
            else:
                seen[(sampleNum, parameterName)] = line

            for column, columnData in enumerate(rowData):
                item = QTableWidgetItem(str(columnData))
                table_widget.setItem(row, column, item)

        layout.addWidget(table_widget)

        size_grip = QSizeGrip(self)
        layout.addWidget(size_grip, 0, Qt.AlignBottom | Qt.AlignRight)

        save_button = QPushButton('Save', self)
        save_button.clicked.connect(self.save_and_close)

        close_button = QPushButton('Close', self)
        close_button.clicked.connect(self.reject)

        layout.addWidget(close_button)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def load_data(self):
        pass;


    def save_and_close(self):
        self.accept()

