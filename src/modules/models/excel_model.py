import os
import math
import platform

from base_logger import logger

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.worksheet.page import PageMargins, PrintOptions

from modules.managers.authors_manager import AuthorsItem

class ExcelModel:

    def __init__(self, jobNum, report_type):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.jobNum = jobNum

        if platform.system() == "Windows":
            self.system = 1
        elif platform.system() == "Darwin":  # MacOS
            self.system = 2

        print(f'self.system : {self.system}')

        #1 = ICP, 2 = CHM
        self.report_type = report_type

        self.total_cols = 8
        self.page_size = 62

        self.font_name = "Times New Roman"
        self.font_size = 9
        self.row_char_limit = 120

        self.default_styles()

    def default_styles(self):
        self.default_font = Font(name="Times New Roman", size=9)
        self.thin_border_style = Side(border_style="thin", color="000000")
        self.double_border_style = Side(border_style='double', color="000000")

        self.thin_right_border = Border(right=self.thin_border_style)
        self.thin_left_border = Border(left=self.thin_border_style)
        self.thin_side_border = Border(right=self.thin_border_style, left=self.thin_border_style)
        self.thin_top_border = Border(top=self.thin_border_style)

        self.centered_alignment = Alignment(horizontal='center', vertical='center')

    def excel_page_setup(self):
        logger.info('Entering excel_page_setup')

        self.ws.sheet_view.view = "pageLayout"

        # Ensure that the scaling happens automatically
        self.ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=False)

        # set custom page margins
        self.ws.page_margins = PageMargins(
            left=0.7,
            right=0.7,
            top=0.75,
            bottom=0.75,
            header=0.3,
            footer=0.3
        )

        # set the excel page settings
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0  # Automatic height fitting
        self.ws.page_setup.orientation = self.ws.ORIENTATION_PORTRAIT

        # set column dimensions
        self.ws.column_dimensions['A'].width = 20 #120px
        self.ws.column_dimensions['H'].width = 20 #1 width is equal to 6 pixels
        self.ws.print_title_rows = '1:8' # the first two row

    def estimate_char_limit_in_cell(self, column_width, font_name, font_size, excel_app=None):
        """
        Estimates the maximum character limit for a cell based on column width, font, and font size.

        Args:
            column_width (float): Width of the column in pixels.
            font_name (str): Name of the font (e.g., "Times New Roman").
            font_size (int): Font size in points.
            excel_app (optional): A running Excel application object (for more accurate results).

        Returns:
            int: Estimated character limit.
        """

        try:
            if excel_app:
                # Use Excel for more accurate measurement (requires running Excel instance)
                xl_sheet = excel_app.Sheets(1)  # Assuming you're working with the first sheet
                xl_sheet.Cells(1, 1).Font.Name = font_name
                xl_sheet.Cells(1, 1).Font.Size = font_size
                xl_sheet.Columns(1).ColumnWidth = column_width / 8.4  # Convert pixels to Excel column width units

                # Test with a representative string
                test_string = "abcdefghijklmnopqrstuvwxyz ABCDEFGHIJKLMNOPQRSTUVWXYZ 0123456789 .,;:!?"
                xl_sheet.Cells(1, 1).Value = test_string

                # Gradually reduce the string length until it fits in one line
                while xl_sheet.Cells(1, 1).WrapText:
                    test_string = test_string[:-1]

                return len(test_string)

            else:
                # **This part is not accurate and is removed**
                # ... (Previous estimation using QLabel was removed)
                # ...

                # Since we don't have a GUI library, we can use a very simple heuristic
                # (This is a rough estimate and might not be very accurate)
                average_char_width = 7  # Adjust this value based on font and font size
                estimated_char_limit = int(column_width / average_char_width)
                return estimated_char_limit


        except Exception as e:
            print(f"Error estimating character limit: {e}")
            return 0

    def format_rows(self, total_samples):
        logger.info(f'Entering format_rows with total_samples: {total_samples}')
        logger.info(f'self.system: {self.system }')

        window_conversion = 0.75
        row_height_pixels = 15

        if(self.system == 1): # Windows
            window_conversion = 0.75
            row_height_pixels = 15
            #TODO: check if the row_height_pixels mess things up lol

        if(self.system == 2): #Mac
            window_conversion = 1
            row_height_pixels = 13

        # extra page just to be safe for the formatting
        total_pages = math.ceil(total_samples/4) + 1
        total_rows = (self.page_size * total_pages) - (8 * (total_pages -1))

        logger.debug(f'total_pages: {total_pages}, total_rows: {total_rows}')

        # iterate through all of the rows and set them the correct height
        for row in self.ws.iter_rows(min_row=1, max_col=self.total_cols, max_row=total_rows):
            for cell in row:
                cell.font = self.default_font
                self.ws.row_dimensions[cell.row].height = (row_height_pixels * window_conversion)

    def set_headers_and_footers(self):
        logger.info('Entering set_headers_and_footers')

        if(self.report_type == 1):
            title = 'CHM REPORT'
        elif(self.report_type == 2):
            title = 'ICP REPORT'
        else:
            title = 'Other REPORT'

        left_text = f"{title}: &D"
        right_text_header = f"Page &P of &N \n W{self.jobNum}"
        right_text_footer = '&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
        center_text_footer = "&B MB Laboratories Ltd.&B \nwww.mblabs.com"
        contact_footer = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com'

        font_size = 10

        # insert the headers
        self.insert_header_footers(self.ws.oddHeader, left_text, right_text_header)
        self.insert_header_footers(self.ws.evenHeader, left_text, right_text_header)

        # insert the footers
        self.ws.oddFooter.font_name = self.font_name
        self.ws.oddFooter.font_size = font_size
        self.ws.evenFooter.font_name = self.font_name
        self.ws.evenFooter.font_size = font_size

        self.insert_header_footers(self.ws.oddFooter, contact_footer, right_text_footer, center_text_footer)
        self.insert_header_footers(self.ws.evenFooter, contact_footer, right_text_footer, center_text_footer)


    def insert_header_footers(self, placement, left_text, right_text, center_text=None):

        font_size = 10

        placement.left.font = self.font_name
        placement.left.size = font_size
        placement.left.text = left_text

        placement.right.font = self.font_name
        placement.right.size = font_size
        placement.right.text = right_text

        if center_text:
            placement.center.font = self.font_name
            placement.center.size = font_size
            placement.center.text = center_text

    def insert_client_info(self, client_info, column2):
        logger.info('Entering insert_client_info')

        if not self.ws:
            raise ValueError("Worksheet is not set. Use set_worksheet() to initialize the worksheet.")

        # Primary client details
        self.ws['A1'] = client_info.get('clientName', '')

        self.ws['A2'] = client_info.get('attn', '*') or '*'
        self.ws['A3'] = client_info.get('addy1', '')
        self.ws['A4'] = f"{client_info.get('addy2', '')}, {client_info.get('addy3', '')}"

        self.ws['A6'] = f"TEL: {client_info.get('tel', '')}"
        self.ws['A7'] = client_info.get('email', '')

        # Secondary client details
        self.ws[f'{column2}1'] = f"Date: {client_info.get('date', '')}  ({client_info.get('time', '')})"
        self.ws[f'{column2}2'] = f"Source: {client_info.get('sampleType1', '')}"
        self.ws[f'{column2}3'] = f"Type: {client_info.get('sampleType2', '')}"
        self.ws[f'{column2}4'] = f"No. of Samples: {client_info.get('totalSamples', '')}"
        self.ws[f'{column2}6'] = f"Arrival temp: {client_info.get('recvTemp', '')}"
        self.ws[f'{column2}7'] = f"PD: {client_info.get('payment', '')}"

        return self.ws

    def format_sample_header_names(self, sample_names):
        ''' combines the sample_name with it's time and date values into one string'''
        logger.info('Entering format_sample_header_names')

        sample_sections = []
        sample_placement = []

        display_limit = 4

        current_word = ''
        temp = []

        for i, (key,value) in enumerate(sample_names.items(), start=1):
            stripped_word = " ".join(value.split())
            current_word += " " + str(i) + ") " + stripped_word + " "
            temp.append(key)

            if(i % display_limit == 0):
                sample_sections.append(current_word)
                sample_placement.append(temp)
                current_word = ""
                temp = []

        if(current_word != ''):
            sample_sections.append(current_word)
            sample_placement.append(temp)

        logger.debug(f'sample_sections: {sample_sections}, sample_placement:{sample_placement}')

        return sample_sections, sample_placement

    def insert_sample_name(self,row, sample_section):
        logger.info(f'Entering insert_sample_name with row: {row}, sample_section: {sample_section}')

        sample_cell = self.ws.cell(row=row, column=1)
        sample_cell.value = f'Samples: {sample_section}'
        sample_cell.border = Border(bottom=self.thin_border_style)

        self.ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=self.total_cols)
        sample_cell.alignment = Alignment(wrap_text=True)

        return row + 3

    def set_cell_value_with_alignment(self, row, col, value, alignment=None):

        if(alignment is None):
            alignment = self.centered_alignment

        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = alignment
        return cell

    def set_cell_value_with_format(self, row, col, value, alignment=None, border=None):

        if(alignment is None):
            alignment = self.centered_alignment

        if(border is None):
            border = self.thin_side_border

        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = alignment
        cell.border = border
        return cell

    def set_sample_names(self, allowed_borders: list, page_location: int, total_samples: int, start_val: int):

        for i in range(total_samples):
            sample = self.ws.cell(row=page_location, column=i + 3)
            allowed_borders.append(i+3)
            sample.value = f"Sample {start_val + i + 1}" if start_val != 0 else f"Sample {i + 1}"
            sample.alignment = self.centered_alignment
            sample.border = self.thin_side_border

        return page_location

    def insert_test_titles(self, page_location: int, total_samples: int, start_val: int):
        logger.info(f'Entering insert_test_titles with page_location: {page_location}, total_samples:{total_samples}, start_val: {start_val}')

        # openpyxl columns are 1-based index
        titles = {}

        if(self.report_type == 1): # ICP Report
            titles = {
                1: 'Elements',
                2: 'Symbols',
                7: 'Units',
                8: 'Maximum Limits'
            }

        if(self.report_type == 2): # CHM Report
            titles = {
                1: 'Tests Name',
                2: 'Units',
                7: 'Lab Blank',
                8: 'Sₒ',
                9: '% Recovery',
                10: 'Maximum Limits',
            }

        start_col = min(titles.keys())
        last_col = max(titles.keys())

        logger.debug(f'titles: {titles}')
        logger.debug(f'start: {start_col}, last: {last_col}')

        for col, col_name in titles.items():

            if(col in [start_col, last_col]):
                self.set_cell_value_with_alignment(page_location, col, col_name , Alignment(horizontal='left', vertical='center', indent=1))

            else:
                cell = self.set_cell_value_with_alignment(page_location, col, col_name)
                cell.border = self.thin_side_border

        # Apply sample names and update page_location
        allowed_borders = list(titles.keys())
        self.set_sample_names(allowed_borders, page_location, total_samples, start_val)

        page_location +=1

        return self.apply_borders_and_formatting(sorted(allowed_borders), page_location)

    def apply_borders_and_formatting(self, allowed_borders, page_location):
        """
        Helper function to apply borders and formatting to rows based on allowed borders.
        """

        for i in range(1, self.total_cols + 1):
            current = self.ws.cell(row=page_location, column=i)
            if i in allowed_borders:
                if i != allowed_borders[-1]:
                    current.border = Border(right=self.thin_border_style, bottom=self.thin_border_style)
                else:
                    current.border = Border(bottom=self.thin_border_style)
                if i == 7:
                    current.border = Border(right=self.thin_border_style, left=self.thin_border_style, bottom=self.thin_border_style)
            else:
                current.border = Border(bottom=self.thin_border_style)

        page_location +=1
        return page_location

    def insert_footer_comment(self, page_location, comment, author):
        page_location = self.insert_comment(page_location, comment)
        page_location +=2

        if(len(author) > 1):
            self.insert_signature(page_location, [3,6], author)
        else:
            self.insert_signature(page_location,[6], author)


    def insert_comment(self, page_location, comment):
        logger.info('Entering insert_comment')

        for _, commentLine in enumerate(comment):
            temp = self.ws.cell(row = page_location, column=1)
            temp.value = commentLine
            page_location+=1

        return page_location

    def insert_next_page_comment(self, page_location):
        logger.info('Entering insert_next_page_comment')
        comment =self. ws.cell(row=page_location, column=1)
        comment.value = 'continued on next page....'
        comment.font = Font(bold=True, size=9, name="Times New Roman")


    def insert_signature(self,  page_location:int , start_col: int, author_info: list) -> int:
        logger.info(f'Entering insert_signature with page_location:{repr(page_location)}, start_col: {repr(start_col)}, author_info: {repr(author_info)}')

        authorNames = []
        authorRoles = []

        for author in author_info:

            if(isinstance(author, AuthorsItem)):
                full_name = author.full_name
                department = author.department
                title = author.title

                department_convert = {
                    'CHEM': 'Analytical Chemist',
                    'MICRO': 'Microbiologist'
                }

                role = 'Sr. ' + department_convert[department] if title == 'S' else department_convert[department]

                authorNames.append(full_name)
                authorRoles.append(role)

        for i, col in enumerate(start_col):
            scientistNamePos = self.ws.cell(row=page_location, column=col)
            scientistRolePos = self.ws.cell(row=page_location+1, column=col)

            scientistNamePos.value = authorNames[i]
            scientistRolePos.value = authorRoles[i]

            for j in range(2):
                signatureLine = self.ws.cell(row=page_location, column=col+j)
                signatureLine.border = Border(top=self.thin_border_style)

    def add_bottom_border(self, page_location):

        for i in range(1, self.total_cols + 1):
            bottomBorder = self.ws.cell(row=page_location, column=i)
            bottomBorder.border = self.thin_top_border

    def save_excel(self, file_path):
        logger.info(f'Entering save_excel with file_path: {file_path}')

        filename, ext = os.path.splitext(file_path)
        counter = 1

        #if os.path.exists(file_path):
            # remove existing file so can write
            # os.remove(file_path)

        while os.path.exists(file_path):
            counter += 1
            file_path = f'{filename}({counter}){ext}'

        if(counter > 1):
            logger.debug('There is already a file with the same name in this location')

        self.wb.save(file_path)

        return file_path


def split_sentence_by_words(sentence, max_chars_per_line=140):
    words = sentence.split()
    lines = []
    current_line = ""

    for word in words:

        if len(current_line) + len(word) + 1 <= max_chars_per_line:  # +1 for the space
            current_line += word + " "
        else:
            lines.append(current_line.strip())  # Add the current line to the list
            current_line = word + " "  # Start a new line with the current word

    if current_line:  # Add the last line if it's not empty
        lines.append(current_line.strip())

    return lines

def significant_figures_convert(value):
    if(value >= 100):
        return (f'{value:.0f}')
    if(value >=10):
        return (f'{value:.1f}')
    if(value >= 1):
        return (f'{value:.2f}')
    if(value < 1):
        return (f'{value:.3f}')

    return value