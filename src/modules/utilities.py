import os 
import re
import csv
import numpy
import sqlite3 
import json 
import string
import pickle 
import openpyxl

from base_logger import logger
from copy import copy
from datetime import date

from PyQt5.QtGui import QColor
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QFileDialog, QPushButton, QTableWidgetItem, 
    QTableWidget , QVBoxLayout, QDialog,  QSizePolicy, QSizeGrip, 
    QGraphicsDropShadowEffect 
)

from modules.constants import *
from modules.dialogBoxes import *

# Project wide reused functions 

#******************************************************************
#   General Functions  
#******************************************************************
def apply_drop_shadow_effect(widget):
    logger.info('Entering apply_drop_shadow_effect')
    # Create drop shadow effect
    shadow = QGraphicsDropShadowEffect(widget)
    shadow.setColor(QColor(0, 0, 0, 150))  # Set shadow color
    shadow.setBlurRadius(10)  # Set shadow blur radius
    shadow.setOffset(0, 0)  # Set shadow offset (no offset)

    logger.info('Applying drop shadow effect to the widget')
    widget.raise_()
    widget.setAutoFillBackground(True) 
    widget.setGraphicsEffect(shadow)


def removeIllegalCharacters(input_string):
    # Define a regular expression pattern to match illegal and escape characters
    pattern = r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]'
    
    # Use re.sub to replace matched characters with an empty string
    cleaned_string = re.sub(pattern, '', input_string)
    
    return cleaned_string

def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


def search_list_of_lists(lists, targets):
    for sublist in lists:
        if all(target in sublist for target in targets):
            return sublist 
    return None

def is_real_number(string): 
    try:
        float(string)
        return True
    except ValueError:
        return False

def remove_unicode_characters(text):
    # Create a translation table with all Unicode characters set to None
    translation_table = dict.fromkeys(range(0x10000), None)
    translation_table.update(str.maketrans("", "", string.printable))

    # Remove Unicode characters using the translation table
    cleaned_text = text.translate(translation_table)

    return cleaned_text

def hardnessCalc(calcium, magnesium, dilution):
    calcium = float(calcium) * float(dilution)
    magnesium = float(magnesium) * float(dilution)
    
    return round(calcium * 2.497 + calcium * 4.11, 1)



#******************************************************************
#    OS Access Functions 
#******************************************************************

def save_pickle(dictionaryName):
    logger.info(f'Entering save_pickle with parameter: dictionaryName: {repr(dictionaryName)}')
    try:
        with open("data.pickle", "wb") as f:
            pickle.dump(dictionaryName, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as ex:
        print("Error during pickling object (Possibly unsupported):", ex)
        logger.error(ex)


def load_pickle(filename):
    logger.info(f'Entering load_pickle with parameter: filename: {repr(filename)}')
    try:
        with open(filename, "rb") as f:
            return pickle.load(f)
    except Exception as ex:
        print("Error during unpickling object (Possibly unsupported):", ex)
        logger.error(ex)
        
 
def openFile(): 
    logger.info('Entering openFile')
    fileName, _ = QFileDialog.getOpenFileName(None, 'Open File', '',)
    return fileName

def getFileLocation():
    logger.info('Entering getFileLocation')
    dlg = QFileDialog().getExistingDirectory()
    print(dlg)
    return dlg

def isValidDatabase(database_path):
    try:
        conn = sqlite3.connect(database_path)
        conn.close()
        return True
    except sqlite3.DatabaseError:
        print
        return False

def saveNewLocation():
    location = getFileLocation() 
    text = input('Save File Name')
    
    save_pickle({text:location})


def scanDir(path): 
    logger.info(f'Entering ScanDir with parameter path: {path}')

    obj = os.scandir(path)
    #file = os.listdir(path)
    
    for entry in obj :
        if entry.is_dir() or entry.is_file():
            print(entry.name)
    

    obj.close()

def scanForTXTFolders(jobNum): 
    logger.info(f'Entering scanForTXTFolders with parameter jobNum: {jobNum}')
    
    fileLocationsDictionary = load_pickle('data.pickle')
    TXTLocation = fileLocationsDictionary['TXTDirLocation']
    locationsObject = os.scandir(TXTLocation)

    txtFolderLocations = [] 
    
    for entry in locationsObject: 
        if(entry.is_dir()):
            if(re.match('^TXT-[a-zA-Z]{3}$', entry.name)):
                txtFolderLocations.append(os.path.join(TXTLocation, entry.name))
            
    locationsObject.close()
    return processTXTFolders(jobNum, txtFolderLocations)
  
 
def processTXTFolders(jobNum, locations):
    logger.info(f'Entering processTXTFolders with parameter jobNum: {jobNum}, locations: {locations}')
    
    fileName = "W" + jobNum + ".TXT"
    
    for i in range(len(locations)): 
        tempLocationObject = os.scandir(locations[i]) 

        for entry in tempLocationObject: 
            if(entry.is_file()): 
                if(re.match(fileName, entry.name)): 
                    logger.info('TXT FILE FOUND')
                    #print(entry.name)
                    tempLocationObject.close()
                    return os.path.join(locations[i], entry.name)
        
        tempLocationObject.close()
    #TODO: return a blank user information 
    #can just clone the clientInfoDict somewhere and send it back 
    #print("No Job Number Matches")
    return None; 


#******************************************************************
#   TEXT File Processing 
#******************************************************************
        
#FIXME: Cannot process big text information 
def processClientInfo(jobNum, fileLocation):
    logger.info(f'Entering processClientInfo with jobNum: {jobNum}')
    
    clientInfoDict = {
        'clientName': '', 
        'date': '', 
        'time': '', 
        'attn': '', 
        'addy1': '', 
        'addy2': '', 
        'addy3': '', 
        'sampleType1': '', 
        'sampleType2': '', 
        'totalSamples': '', 
        'recvTemp': '', 
        'tel': '', 
        'email': '', 
        'fax': '', 
        'payment': ''
    }
    
    #grab the file names 
    sampleNames = {}
    
    #have the information about the file, what kind of reports and etc 
    sampleTests = {}

    sampleCounter = 0; 
    prevLine = [0, ""]
    prevLineHelper = [0, ""]
    
    if(fileLocation == None): 
        logger.info('Completed Processing Client Info returning info')
        logger.info(f'Client Information Dictionary: {clientInfoDict}')
        logger.info(f'Sample Names: {repr(sampleName)}')
        logger.info(f'Sample Tests: {sampleTests}')
        
        return clientInfoDict, sampleNames, sampleTests; 
    
    with open(fileLocation) as file: 
        logger.debug(f'Opening File: {repr(fileLocation)}')    
        sampleNum = ''; 
    
        for lineLocation, line in enumerate(file, 0):
            
            logger.debug(f'{lineLocation}: {repr(line)}')
            
            #TODO: can put this into a own function 
            # Scan the initial 8 lines for information
            
            if(prevLine[0]+1 == prevLineHelper[0]):
                prevLine[0] = copy(prevLineHelper[0])
                prevLine[1] = copy(prevLineHelper[1])
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
            else: 
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
                    
            if(lineLocation == 1): 
                clientInfoDict['clientName'] = line[0:54].strip()
                clientInfoDict['date'] = line[50:(54+7)].strip()
                clientInfoDict['time'] = line[66:71].strip()
                   
            if lineLocation == 2:
                clientInfoDict['sampleType1'] = line[54:].strip()
            
                if "*" in line: 
                    clientInfoDict['attn'] = line[:54].strip()
                    
                else: 
                    clientInfoDict['addy1'] = line[:54].strip()
                
            if(lineLocation == 3): 
                clientInfoDict['sampleType2'] = line[54:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy1'] = line[:60].strip()
                else: 
                    clientInfoDict['addy2'] = line[:60].strip()
            
            if(lineLocation == 4): 
                clientInfoDict['totalSamples'] = line[60:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy2'] = line[:60].strip()
                else: 
                    clientInfoDict['addy3'] = line[:60].strip() 
                    
            if(lineLocation == 5): 
                if(clientInfoDict['attn'] and clientInfoDict['addy2']): 
                    clientInfoDict['addy3'] = line[:60].strip()
                else: 
                    clientInfoDict['tel'] = line[26:50].strip()

                    try: 
                        clientInfoDict['recvTemp'] = line[71:].strip()
                    except:
                        logger.debug('No recv temp available')
                    
                        
            if(lineLocation == 6): 
                clientInfoDict['tel'] = line[26:50].strip() 
                clientInfoDict['recvTemp'] = line[71:].strip()
            
            if(lineLocation == 7): 
                clientInfoDict['fax'] = line[26:].strip()
                
            if(lineLocation == 8): 
                
                try: 
                    foundEmail = re.search('([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+', line).group()
                    if(foundEmail): 
                        clientInfoDict['email'] = foundEmail; 
                except:
                    logger.error('Email Error')
                
                
                if("pd" in line.lower()): 
                    clientInfoDict['payment'] = line[51:].strip()
                     
            #FIXME: error with 172235 ICP, not showing the proper text amount
            #FIXME: 
            # Scan the bottom half to  lines for information 
            if(lineLocation > 35 and len(line) > 0): 
                #FIXME: don't think I need this try-except section 
                if(sampleCounter != int(clientInfoDict['totalSamples']) ):
                    try: 
                        sampleNumCheck = re.search(r"\d+", line)
                        sampleNameCheck = re.search('(?<=\s[0-9]).*', line)
                                                
                        if(sampleNameCheck): 
                            sampleNum = sampleNumCheck.group() 
                            sampleName = sampleNameCheck.group().strip() 

                            #print(f'**Sample Name: {sampleName} Sample Num: {sampleNum}')
                            logger.info(f'**Sample Name: {repr(sampleName)} Sample Num: {repr(sampleNum)}')
                            
                            #combined_name = str(jobNum) + '-' + str(sampleCounter+1)
                            combined_name = str(jobNum) + '-' + str(sampleNum)
                            sampleNames[combined_name] = sampleName
                            #sampleCounter+=1; 
                           
                        #TODO: add something to get the - afterwords names
                    except Exception as error: 
                        print(f'[ERROR]: {error}')
                        print('Invalid Line: Name Assignment Fail')

                #find the report information that does with corresponding thing                
                if(re.search('(?<=\s-\s).*', line)):
                    #prevSampleName = str(jobNum) + "-" + str(sampleCounter-1)
                    prevSampleJobName = str(jobNum) + '-' + str(sampleNum)
                    
                    currentLineTestsCheck = re.search('(?<=\s-\s).*', line)
                    prevSampleMatchCheck = re.search('(?<=\s[0-9]).*', prevLine[1])
                    prevSampleTestsCheck = re.search('(?<=\s-\s).*', prevLine[1])
                    #sampleTests[prevSampleName] = currentTestsCheck.group()
                    #not could be apart of the string name longer 
                    
                    #add to most recent sample
                    if(currentLineTestsCheck):
                        logger.debug('currentLineTestsCheck: True')
                        
                        print(f'**Prev Sample Name: {prevSampleJobName} Sample Name: {sampleName} Sample Num: {sampleNum}')
                        #print(f'Current is a test: {currentLineTestsCheck.group()}')
                        #print(f'SAMPLE TESTS: {sampleTests}')
 
                        if(prevSampleTestsCheck):
                            logger.debug('currentLineTestsCheck & prevSampleTestsCheck: True')
                            sampleTests[prevSampleJobName] = sampleTests[prevSampleJobName]  + ", " + currentLineTestsCheck.group()
                        else: 
                            logger.debug('currentLineTestsCheck & prevSampleTestsCheck: False')
                            sampleTests[prevSampleJobName] = currentLineTestsCheck.group()
                        
                    #Prev sample name 
                    if(prevSampleMatchCheck):
                        logger.debug('prevSampleMatchCheck: True')
                        #print('prev was a sampleName')
                        #print(sampleName[prevSampleName]) #doesn't work 
                        pass; 
                       
                    #append onto them 
                        
                    #TODO: solve this later, add previous name onto current name sample 
                    if((not bool(prevSampleMatchCheck)) and not( bool(prevSampleTestsCheck))): 
                        logger.debug('prevSampleMatchCheck: False and prevSampleTestsCheck: False') 
                        

    file.close()

    #process type sampleTests 
    for key,value in sampleTests.items():
        testLists = [x.strip() for x in value.split(',')]
        sampleTests[key] = testLists
            
    logger.info('Exiting processClientInfo and returning')
    logger.info(f'Client Information Dictionary: ')
   
    for key, value in clientInfoDict.items(): 
        logger.info(f'*{key}: {repr(value)}')
    
    logger.info(f'Sample Names: {repr(sampleName)}')
    logger.info(f'Sample Tests: {sampleTests}')
    
    return clientInfoDict, sampleNames, sampleTests; 



#TODO: move this into icp tools 
#******************************************************************
#    ICP Uploads Files 
#******************************************************************
    
def icp_upload(filePath, db): 
    logger.info('Entering icp_upload with filePath: {filePath}')

    logger.info('Scanning File... ')
    if(filePath.endswith('.txt')):
        icpMethod1(filePath, db)
    elif(filePath.endswith('.xlsx')):
        icpMethod2(filePath, db)
    else: 
        print("Not valid file type")
    return; 
    
#TODO: sort by name  
#read line by line and just add the line instead 
#FIXME: tracking the icpData is wrong as well 
 #TODO: insert try catch block 
def icpMethod1(filePath, db): 
    logger.info('Entering icpMethod1 txt filetype detected')

    file1 = open(filePath, 'r')
    baseName = os.path.basename(filePath)
    fname = baseName.split('.txt')[0]
    #remove extension 

    logger.debug('FileName: ', fname)
    
    Lines = file1.readlines()
    
    startingLine = 'Date Time Label Element Label (nm) Conc %RSD Unadjusted Conc Intensity %RSD' 
    headers = ['Sample', 'Analyze', 'Element', 'HT', ' ', 'units', 'rep', ' ', ' '] 
    
    startingPostion = []
    endPostion = []
    count = 0
    
    # Strips the newline character
    for line in Lines:
        
        #print("Line{}: {}".format(count, line.strip()))
        if(line.strip() == startingLine):
            startingPostion.append(count)
            
        if(re.search('([1-9]|[1-9][0-9]) of ([1-9]|[1-9][0-9])$', line)): 
            endPostion.append(count)

        count += 1

    #update headers 
    headerUpdate = Lines[startingPostion[0] + 1]; 
    headerUpdate = headerUpdate.split()
    headers[7] = headerUpdate[0]
    headers[8] = headerUpdate[1]

    newName = fname + '.csv'
    loadPath = load_pickle('data.pickle') 
    newPath = os.path.join(copy(loadPath['ispDataUploadPath']), newName) 
     
    logger.info('Writing CSV File: {}'.format(newPath))
    
    f = open(newPath, 'w')
    writer = csv.writer(f)
    writer.writerow(headers)
    
    spiltLengths = []
    
    jobNumbers = []
    jobData = {}
    
    elementData = {}
    currentJob = ''

    sampleNumbers = []
    
    #TODO: missing the last row 
    for start in startingPostion: 
        running = True; 
        counter = 1; 

        while(running): 

            currentLine = Lines[start + counter]
            
            if(re.search('([1-9]|[1-9][0-9]) of ([1-9]|[1-9][0-9])$', currentLine)): 
                break; 
            
            counter += 1; 

            splitLine = currentLine.split()
            spiltLengths.append(len(splitLine))

            if(re.search('\d{6}-\d{1,2}', splitLine[2])):
                temp = []
                temp.append(splitLine[2])
                temp.append(1)
                temp.append(splitLine[3])
                temp.append(1)
                temp.append(splitLine[6])
                temp.append('mg/L')
                temp.append(1)
                temp.append(splitLine[0])
                temp.append(splitLine[1])

                sampleNumbers.append([splitLine[2], start+counter, splitLine[3], splitLine[6]])

                if(currentJob == ''):
                    currentJob = splitLine[2]
  
                elif(currentJob != splitLine[2]): 
                    jobData[currentJob] = elementData   
                    elementData = {}
                    currentJob = splitLine[2]

                elementData[splitLine[3]] = splitLine[6]
    
                jobNumber = splitLine[2].split('-')[0]
                if(jobNumber not in jobNumbers):
                    jobNumbers.append(jobNumber)
                    
                if(temp): 
                    writer.writerow(temp)

        jobData[currentJob] = elementData   

    spiltLengths = numpy.array(spiltLengths)
    unique, counts = numpy.unique(spiltLengths, return_counts=True)


    f.close()
    file1.close()
    
    #TODO: factor in a way to show all the things and the lines where duplicates are 
    logger.debug(f'SampleNumbers: {sampleNumbers}')
    logger.debug(f'JobData: {jobData}')

    save = viewIcpTable(filePath, sampleNumbers, 1) 
    logger.info(f"Save Status: {repr(save)}")


    #save to database 
    if(save): 
        for (key, value) in jobData.items(): 
            #
            #sql = 'INSERT OR REPLACE INTO icpMachineData1 values(?,?,?,?,?, 1)'
            sql = 'INSERT OR REPLACE INTO icpData values(?, ?, ?, ?, ?, 1)'
            
            jobNum = key.split('-')[0]
            todayDate = date.today()
            
            tempData = json.dumps(value)
            if(key): 
                db.execute(sql, (key,jobNum,baseName, tempData, todayDate))
                db.commit()
        return jobNumber,jobData
    else: 
        return False; 

    
#scans thought all the text files and finds all the different sample types and the ISP and CHM files     
#TODO: insert try catch block 
def icpMethod2(filePath, db): 
    logger.info('Entering icpMethod1 xlsx filetype detected')
    wb = openpyxl.load_workbook(filePath)
    sheets = wb.sheetnames 
    
    baseName = os.path.basename(filePath)
    fname = baseName.split('.xlsx')[0]
    
    print('Method 2')
    print('FileName: ', fname)

    newName = fname + '_formated'  + '.xlsx'
    loadPath = load_pickle('data.pickle') 
    newPath = os.path.join(copy(loadPath['ispDataUploadPath']), newName)  
    
    ws = wb[sheets[0]]

    sampleTypeColumn = ws['E']
    sampleNameColumn = ws['G']
    
    elementConversion = ['As', 'Se', 'Cd', 'Sb', 'Hg', 'Pb', 'U']
    elementColumns = ['I', 'J', 'M', "Q", 'U', 'AA', 'AC']
    selectedRows = []

    sampleInfo = {}

    for cell in sampleTypeColumn:     
        if(cell.value == 'Sample'):
             
            currentSampleName = ws.cell(row=cell.row, column=7).value
            pattern = r'^\d{6}-\d{3}'
            
            if(re.search(pattern, currentSampleName)): 
                selectedRows.append(cell.row)
                
                sampleName = formatJobSampleString(currentSampleName)
                jobNum = sampleName[:6]
                #print('---------------')
                #print(sampleName, '|' , jobNum)

                sampleData = {}
                
                for i, element in enumerate(elementColumns): 
        
                    col_index = openpyxl.utils.column_index_from_string(element)
                    elementVal = ws.cell(row=cell.row, column=col_index).value 

                    if(elementVal == '<0.000'):
                        sampleData[elementConversion[i]] = 0.00
                    else: 
                        sampleData[elementConversion[i]] = elementVal
                    
                sampleInfo[sampleName] = sampleData
                
    for key, value in sampleInfo.items(): 
        query  = 'INSERT OR REPLACE INTO icpData values(?, ?, ?, ?, ?, 2)'
        jobNum = key[:6]
        todayDate = date.today()
        tempData = json.dumps(value)
        if(jobNum): 
            db.execute(query, (key, jobNum, baseName, tempData, todayDate))
            db.commit()

    # Format the machine data 
    formatMachineData(ws, selectedRows, elementColumns, newPath)
    
    return; 

def formatMachineData(ws, selectedRows, elementColumns, newPath): 
    logger.info('Entering formatMachineData with parameters:')
    logger.info(f'selectedRows: {repr(selectedRows)}')
    logger.info(f'elementColumns: {repr(elementColumns)}')
    logger.info(f'newPath: {repr(newPath)}')

    newWb = openpyxl.Workbook()
    ws2 = newWb.active 
    
    ws2.cell(row=1, column=1).value = 'Sample'
    tableNames = ['', 'Rjct', 'Data File', 'Acq. Date-Time', 'Type', 'Level', 'Sample Name', 'Total Dil', 
                  '75  As  [ He ] ', '78  Se  [ H2 ] ', '111  Cd  [ No Gas ] ', '123 Sb [He]', '202 Hg [He]', '208 Pb [He]', '238 U[He]'
                  ]
    
    ws2.merge_cells('A1:H1')
    
    column_num = 1;
    for item in tableNames: 
        ws2.cell(row=2, column=column_num).value = item; 
        column_num += 1;  
   
    row_num = 3; 
    for row_value in selectedRows:

        for currentPos in range(1,8): 
            ws2.cell(row=row_num, column=currentPos).value = ws.cell(row=row_value, column=currentPos).value
        
        tempCol = 9
        for col in elementColumns: 
            col_index = openpyxl.utils.column_index_from_string(col)
            ws2.cell(row=row_num, column=tempCol).value = ws.cell(row=row_value, column=col_index).value
            tempCol+=1; 
            

        row_num+=1;      

    newWb.save(newPath)

def formatJobSampleString(inputString): 
    sample = inputString.strip()
    match = re.match(r'(\d+)-0+(\d+)', sample)
    
    if match: 
        first_part = match.group(1)
        second_part = match.group(2)
        formatted_string = f"{first_part}-{second_part}"
        
        return formatted_string; 

def formatStringArray(inputArray): 
    outputArray = []
    
    for sample in inputArray: 
        sample = sample.strip()
        
        # Use regular expressions to extract the desired pattern
        match = re.match(r'(\d+)-0+(\d+)', string)
        
        if match:
            # Get the captured groups from the regex match
            first_part = match.group(1)
            second_part = match.group(2)

            # Construct the desired format
            formatted_string = f"{first_part}-{second_part}"
            outputArray.append(formatted_string)
    
    return(outputArray)
        


    
def viewIcpTable(filePath, data, reportType): 
    logger.info(f'Entering viewIcpTable with parameters: filePath: {filePath}, data: {data}, reportType: {reportType}')

    dialog = icpTableView(filePath, data, reportType)
    
    if dialog.exec_() == QDialog.Accepted:
        return True; 
    else:
        return False; 

#******************************************************************
#    Classes 
#******************************************************************
class icpTableView(QDialog):
    def __init__(self, filePath, data, reportType):
        super().__init__()
        
        self.resize(600, 500)
        self.setWindowTitle(filePath)
    
        headers = ['Sample Number', 'Line', 'Element', 'Value', 'duplicate Line'] 

        layout = QVBoxLayout()
        
        table_widget = QTableWidget()
        table_widget.setRowCount(len(data))
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)
    
        seen = {}
        
        for row, rowData in enumerate(data):
            sampleNum = rowData[0]
            line = rowData[1]
            parameterName = rowData[2]
            
            if (sampleNum, parameterName) in seen: 
                previousRow = seen[(sampleNum, parameterName)] 
                item = QTableWidgetItem(str(previousRow))
                table_widget.setItem(row,4, item)
            else: 
                seen[(sampleNum, parameterName)] = line
            
            for column, columnData in enumerate(rowData):
                item = QTableWidgetItem(str(columnData))
                table_widget.setItem(row, column, item)
                
        layout.addWidget(table_widget)
        
        size_grip = QSizeGrip(self)
        layout.addWidget(size_grip, 0, Qt.AlignBottom | Qt.AlignRight)
        
        save_button = QPushButton('Save', self)
        save_button.clicked.connect(self.save_and_close)
        
        close_button = QPushButton('Close', self)
        close_button.clicked.connect(self.reject)
        
        layout.addWidget(close_button)
        layout.addWidget(save_button)

        self.setLayout(layout)

        
    def save_and_close(self):
        self.accept()

            

