import math 

from base_logger import logger

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.worksheet.header_footer import HeaderFooter

#define default values 
pageRows = 46
defaultFont = Font(name="Times New Roman", size=9)

#Borders 
thinBorder = Side(border_style="thin", color="000000")
doubleBorder = Side(border_style='double', color="000000")


#******************************************************************
#   Excel Document Formatting
#******************************************************************

def pageSetup(ws): 
    logger.info('Entering pageSetup')
    # set the default view to page layout
    ws.sheet_view.view = "pageLayout" 

    logger.info('Preparing to set page width to auto')
    page_setup = ws.page_setup
    
    logger.info('Preparing to set page size defaults')
    page_setup.fitToPage = True
    page_setup.fitToHeight = False 
    page_setup.fitToWidth = True 
    
    logger.info('Preparing to set page margins')
    page_margins = PageMargins()
    page_margins.left = 0.7
    page_margins.right = 0.7
    page_margins.top = 0.75
    page_margins.bottom = 0.75
    page_margins.header = 0.3
    page_margins.footer = 0.3
    
    ws.page_margins = page_margins
    
#FORMATS WAY MORE THEN NEEDED BUT W/
def formatRows(ws, totalSamples, totalCols, pageSize): 
    logger.info(f'Entering formatRows with parameters: totalSamples: {repr(totalSamples)}, totalCols: {repr(totalCols)}, pageSize: {(repr(pageSize))}')

    #WINDOWS VALUES 
    #window_conversion = 0.75 
    #row_height_pixels = 15
    
    #MAC OS VALUES 
    window_conversion = 1 
    row_height_pixels = 13
    
    totalPages = math.ceil(totalSamples/4) 
    totalRows = (pageSize * totalPages) - (8 * (totalPages-1))
    
    logger.debug(f'Total Pages: {repr(totalPages)}')
    logger.debug(f'Total Rows : {repr(totalRows)}')

    for row in ws.iter_rows(min_row=1, max_col=totalCols, max_row=totalRows): 
        for cell in row:
            cell.font = defaultFont 
            ws.row_dimensions[cell.row].height = (row_height_pixels * window_conversion)
            
#******************************************************************
#   General Functions
#******************************************************************
def generateSampleHeaderNames(ws, sampleNames): 
    logger.info(f'Entering generateSampleHeaderNames with parameters: {sampleNames}')

    displayLimit = 4 
    
    sampleSections = []
    samplePlacement = []
    
    currentWord = ""
    temp = []

    for i, (key,value) in enumerate(sampleNames.items(), start=1): 
        stripedWord = " ".join(value.split())
        currentWord += " " + str(i) + ") " + stripedWord + " "
        temp.append(key)

        if(i % displayLimit == 0): 
            print(currentWord)
            sampleSections.append(currentWord)
            samplePlacement.append(temp)
            currentWord = ""
            temp = []
    
    if(currentWord != ''):
        sampleSections.append(currentWord)
        samplePlacement.append(temp) 

    logger.debug(f'Returning the values')
    logger.debug(f'Sample Placement: {samplePlacement}')
    logger.debug(f'Sample Sections: {sampleSections}')

    return sampleSections, samplePlacement 

#******************************************************************
#   Excel Insertion Functions 
#******************************************************************

def createFooters(ws, title, jobNumber): 
    logger.info(f'Entering createFooters with parameters: title:{title}, jobNumber: {jobNumber}')
    
    ws.oddHeader.left.font = 'Times New Roman'
    ws.oddHeader.left.size = 11
    ws.oddHeader.right.font = 'Times New Roman'
    ws.oddHeader.right.size = 11

    ws.evenHeader.left.font = 'Times New Roman'
    ws.evenHeader.left.size = 11
    ws.evenHeader.left.font = 'Times New Roman'
    ws.evenHeader.left.size = 11
    
    ws.oddHeader.left.text  = title + ': &D'
    ws.evenHeader.left.text = title + ': &D' 
    
    ws.oddHeader.right.text  = f"Page &P of &N \n W{jobNumber}"
    ws.evenFooter.right.text = f"Page &P of &N \n W{jobNumber}"
    
    ws.oddFooter.left.text  = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com'
    ws.evenFooter.left.text = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com' 
    
    ws.oddFooter.center.text = "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    ws.evenFooter.center.text= "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    
    ws.oddFooter.right.text = '&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    ws.evenFooter.right.text ='&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    

def insertClientInfo(ws, clientInfo, column2): 
    logger.info('Entering createHeader')
    
    ws['A1'] = clientInfo['clientName']
        
    if(clientInfo['attn'] is None): 
        ws['A2'] = '*' 
    else: 
        ws['A2'] = clientInfo['attn'] 
        
    ws['A3'] = clientInfo['addy1']
    ws['A4'] = clientInfo['addy2'] + ", " + clientInfo['addy3']
    
    ws['A6'] = 'TEL' + clientInfo['tel']
    ws['A7'] =  clientInfo['email']
    
    ws[column2 + '1'] = "Date: " + clientInfo['date'] + "  (" + clientInfo['time'] + ")" 
    ws[column2 + '2'] = "Source: " + clientInfo['sampleType1']
    ws[column2 + '3'] = "Type: " + clientInfo['sampleType2']
    ws[column2 + '4'] = "No. of Samples: " + clientInfo['totalSamples']
    ws[column2 + '6'] = "Arrival temp: " + clientInfo['recvTemp']
    ws[column2 + '7'] = "PD: " + clientInfo['payment']
    
    return ws 
     

def insertSampleName(ws, row, sampleSection, totalRows): 
    logger.info(f'Entering insertSampleName with parameters: row: {row}, totalRows: {totalRows}, sampleSection: {sampleSection}')
    temp = ws.cell(row=row, column=1)
    temp.value = 'Samples: ' + sampleSection
    temp.border = Border(bottom=thinBorder)
   
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=totalRows)
    temp.alignment = Alignment(wrap_text=True) 
    
    return row + 3; 
    
def insertTestTitles(ws, pageLocation, totalSamples, startVal, reportType, totalCols): 
    logger.info(f'Entering insertTestTitles with parameters: pageLocation: {pageLocation}, totalSamples: {totalSamples}, startVal: {startVal}, totalCols: {totalCols}')
    
    tests = ws.cell(row=pageLocation, column=1)
    if(reportType == 0):
        tests.value = 'Tests'
    else: 
        tests.value = 'Elements'        

    units = ws.cell(row=pageLocation, column=2)
    if(reportType == 0): 
        units.value = 'Units'
    else: 
        units.value = "Symbols"

    units.border = Border(right=thinBorder, left=thinBorder)
    units.alignment = Alignment(horizontal='center', vertical='center')
    
    allowedBorders = [1,2,7,8]
    
    #assign sample names 
    for i in range(totalSamples): 
        sample = ws.cell(row=pageLocation, column=i+3)
        allowedBorders.append(i+3)
        
        if(startVal == 0): 
            sample.value = 'Sample ' + str(i + 1)
        else: 
            sample.value = 'Sample ' + str(startVal + i +1) 
             
        sample.alignment = Alignment(horizontal='center', vertical='center')
        sample.border = Border(right=thinBorder, left=thinBorder) 

    so = ws.cell(row=pageLocation, column=7) 
    if(reportType == 0): 
        so.value = 'Recovery'
    else: 
        so.value = 'Units'
    so.border = Border(right=thinBorder, left=thinBorder)
    so.alignment = Alignment(horizontal='center', vertical='center') 
    
    
    colH = ws.cell(row=pageLocation, column=8)
    if(reportType == 0 ): 
        colH.value = 'Comment'
        colH.alignment = Alignment(horizontal='center', vertical='center')  
        
    else:
        colH.value = 'Maximum Limits'  
        colH.alignment = Alignment(horizontal='left', vertical='center', indent=1)  
    
    pageLocation+=1; 
    
    for i in range(1,totalCols+1): 
        current = ws.cell(row=pageLocation, column=i) 
        
        if(i in allowedBorders): 
        
            if(i != 8): 
                current.border = Border(right=thinBorder, bottom=doubleBorder)
            else: 
                current.border = Border(bottom=doubleBorder) 
            
            if(i == 7): 
                current.border = Border(right=thinBorder, left=thinBorder, bottom=doubleBorder)
            
        else: 
            current.border = Border(bottom=doubleBorder)
                     
    pageLocation+=1; 

    return pageLocation 

#will change how we get the tests and unit type, will be saved somewhere 
#TODO: better ways to pass function instead of a large amount of parameters 
def insertTestInfo(ws, pageLocation, testInfo, samplePlacement, sampleData, totalTests, unitType, recovery): 
    logger.info('Entering insertTestInfo')    
    counter = pageLocation

    for i in range(len(testInfo)): 
        #is this removing the escape characters? 
        #currentTest = re.sub('[^A-Za-z0-9]+', '', testInfo[i])
        testPlacement = ws.cell(row=counter, column=1)
        unitPlacement = ws.cell(row=counter, column=2)
        recoveryPlacement = ws.cell(row=counter, column=7)

        try:  
            testPlacement.value = testInfo[i]
        except: 
            testPlacement.value = 'Error'
            
        unitPlacement.value = unitType[i]
        recoveryPlacement.value = recovery[i]
        
        #testPlacement.border = Border(right=thinBorder) 
        unitPlacement.border = Border(right=thinBorder, left=thinBorder)
        unitPlacement.alignment = Alignment(horizontal='center', vertical='center')
        
        recoveryPlacement.border = Border(left=thinBorder, right=thinBorder)
        recoveryPlacement.alignment = Alignment(horizontal='center', vertical='center')  
    
        counter+=1; 
    
    counter = pageLocation 
    
    #insert test user information
    for i, sample in enumerate(samplePlacement, start=3): 
        
        currentResults = sampleData[sample]
        
        for j in range(0,totalTests): 
            currentSample = ws.cell(row=counter+j,column=i)
            currentSample.alignment = Alignment(horizontal='center', vertical='center')
            
            if(type(currentResults[j]) == (int or float)): 
               
                currentSample.value = float(currentResults[j])
            else:
                currentSample.value = currentResults[j]
            
            currentSample.value = currentResults[j]
            currentSample.border = Border(right=thinBorder, left=thinBorder)
                     
    pageLocation += totalTests; 
    
    for i in range(1,9): 
        ws.cell(row=pageLocation, column=i).border = Border(top=thinBorder)
            
    pageLocation += 1; 
    
    return pageLocation; 


def insertComment(ws, pageLocation, comment): 
    logger.info(f'Entering insertComments with parameters: pageLocation: {repr(pageLocation)}, comment: {repr(comment)}')
    
    for i, commentLine in enumerate(comment): 
        temp = ws.cell(row = pageLocation, column=1)
        temp.value = commentLine 
        pageLocation+=1; 

    return pageLocation

def insertNextPageComment(ws, pageLocation): 
    logger.info('Preparing to insert continued to next page...')
    comment = ws.cell(row=pageLocation, column=1)
    comment.value = 'continued on next page....'
    comment.font = Font(bold=True, size=9, name="Times New Roman")  
   
   
def insertSignature(ws, pageLocation:int , startColumn: int, authorsInto: list) -> int: 
    logger.info(f'Entering insertSignature2 with param: pageLocation:{repr(pageLocation)}, startColum: {repr(startColumn)}, authorsInfo: {repr(authorsInto)}')
    
    authorNames = []
    authorRoles = []
    
    for author in authorsInto: 
        authorNames.append(author[1])
        authorRoles.append(author[2])
        
    for i, col in enumerate(startColumn): 
        scientistNamePos = ws.cell(row=pageLocation, column=col)
        scientistRolePos = ws.cell(row=pageLocation+1, column=col)
        
        scientistNamePos.value = authorNames[i]
        scientistRolePos.value = authorRoles[i]
        
        for j in range(2): 
            signatureLine = ws.cell(row=pageLocation, column=col+j)
            signatureLine.border = Border(top=thinBorder)

        
def significantFiguresConvert(value): 
    if(value >= 100): 
        return (f'{value:.0f}')
    if(value >=10): 
        return (f'{value:.1f}')
    if(value >= 1): 
        return (f'{value:.2f}')
    if(value < 1): 
        return (f'{value:.3f}') 
    
    return value; 

#TODO: move this into logic
def floatCheck(s):
    try:
        float(s)
        return True
    except ValueError:
        return False