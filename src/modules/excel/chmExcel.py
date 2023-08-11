import math 
import os 

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.worksheet.header_footer import HeaderFooter

from modules.utilities import * 
from modules.constants import *
from modules.createExcel import *

def createChmReport(clientInfo, jobNum, sampleNames, sampleData, testInfo, unitType, recovery):
    print('***CREATING CHM REPORT')
    print(clientInfo)
    print(sampleNames)
    print(sampleData)
    print(testInfo)
    print(unitType)
    print(recovery)
    print("--------------------")
   
    #FIXME: picke saving and loading without being fucked up, get path of where things should be saved
    temp = load_pickle('data.pickle') 
    exportPath = temp['reportsPath']
    
    wb = Workbook()
    ws = wb.active
    
    pageSetup(ws); 
    
    reportTitle = 'CHM REPORT'
    createFooters(ws, reportTitle, jobNum); 

    #SETTING UP MAIN INFORMATION 
    #TODO: determine how far out the longest thing is 
    #FIXME: check out for none issue when they appear
    
    ws = createHeader(ws, clientInfo, 'D')
    ws.column_dimensions['A'].width = 20 #120px 
    ws.column_dimensions['H'].width = 19
    ws.print_title_rows = '1:8' # the first two rows
    
    totalCols = 8 
    pageSize = 61 
 
    sampleSections = []
    samplePlacement = []
    currentWord = ''
    temp = []
    
    totalSamples = len(sampleData.keys())
    print('Total Samples: ', totalSamples)
    #FIXME: so page amounts change based on column width, so takes in how many pages we want 
    formatRows(ws, totalSamples, totalCols, pageSize);
    
    #can only display 4 at a time 
    currentWord = ""
    temp = []
    
    sampleSections, samplePlacement = generateSampleHeaderNames(ws, sampleNames)
    
    print('Sample Placement: ', samplePlacement)
    print('Sample Sections: ', sampleSections)
    
    pageSize = 56; 
    allocatedSpace = 20; 
    
    testSize = len(testInfo)
    tableSize = 6 + testSize 
    
    totalSampleSections = len(sampleSections)
    totalTablesWithComments = math.floor((pageSize - allocatedSpace)/tableSize)
    totalPages = math.ceil(totalSampleSections/totalTablesWithComments)
    
    print('Total Sample Sections: ', totalSampleSections)
    print('Tables with comments: ' , totalTablesWithComments)
    print('Total Pages: ', totalPages)
    
    usedSamples = 0; 
    pageLocation = 9; 
    totalTests = len(testInfo)

    #determine when the next page is? 
    counter = 0; 
    
    for currentPage in range(totalPages): 
        print('current page:', currentPage)
        sampleAmount = len(samplePlacement[counter])
  
        if(currentPage != 0): 
            pageLocation = (pageSize * currentPage) - (8 * (currentPage-1)) + 1  
        else: 
            pageLocation = 9; 
        
        #last page 
        if(currentPage+1 == totalPages): 
 
            remainingSamples = totalSampleSections - counter; 
            
            for i in range(remainingSamples): 
                pageLocation = insertSampleName(ws, pageLocation, sampleSections[counter], totalCols)
                pageLocation = insertTestTitles(ws,pageLocation, sampleAmount, usedSamples, 0, totalCols)
                pageLocation = insertTestInfo(ws,pageLocation, testInfo, samplePlacement[counter], sampleData, totalTests, unitType, recovery)
                
                if(i+1 == remainingSamples): 
                    pageLocation = insertComments(ws, pageLocation)  
                    pageLocation+=2; 
                    insertSignature(ws, pageLocation, [4,7])
                        
                counter+=1; 
                usedSamples += 4;
        
        #not the last page
        else: 
            print('Not Last Page')
            for i in range(totalTablesWithComments): 
                pageLocation = insertSampleName(ws, pageLocation, sampleSections[counter], totalCols)
                pageLocation = insertTestTitles(ws,pageLocation, sampleAmount, usedSamples, 0, totalCols)
                pageLocation = insertTestInfo(ws,pageLocation, testInfo, samplePlacement[counter], sampleData, totalTests, unitType, recovery)   
                
                if(i+1 == totalTablesWithComments): 
                    comment = ws.cell(row=pageLocation, column=1)
                    comment.value = 'continued on next page....'
                    comment.font = Font(bold=True, size=9, name="Times New Roman")  
                    
                counter+=1; 
                usedSamples += 4; 
        
    
    fileName = 'W' + str(jobNum) + ".chm" 
    filePath = os.path.join(exportPath, fileName)
    print('Export Path: ', filePath)
    
    wb.save(filePath)
    