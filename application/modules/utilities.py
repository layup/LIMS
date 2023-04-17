
import sys
import os 
import re
import csv
import numpy
import inquirer
import json 
import pickle 
from copy import copy
import pandas as pd 
import openpyxl
from datetime import date
from string import printable


from PyQt5.QtWidgets import (
    QFileDialog
)

REPORTS_TYPE = ['','ISP-1','ISP-2','GSMS']
MATRIX_TYPE = ['','food', 'water', 'apples']

PARAMETERS_TYPE = {}
METHODS_TYPE = {}

GSMS_values = {
    "001": ["ALkalinity", "mg/L"], 
    "002": ["NH3-N", "ug/L "], 
    "003": ["Cl-", 'mg/L'], 
    "004": ['E.C', 'us/cm'], 
    "005": ['F-', 'mg/L'], 
    "006": ['TKN', 'mg/L'], 
    "007": ['Mn', 'mg/L'], 
    "008": ['NO3-N', 'ug/L'], 
    "009": ['NO2-N', 'ug/L'], 
    "010": ['ortho-PO43', 'ug/L'],
    "011": ['pH', ' '], 
    "012": ['TPO43 --P', 'ug/L'], 
    "013": ['D.TO43 --P', 'ug/L'], 
    "014": ['SO42', 'mg/L'], 
    "015": ['T.O.C', 'mg/L'], 
    "016": ['T&L', 'mg/L'],
    "017": ['TDS', 'mg/L'], 
    "018": ['TSS', 'mg/L'],
    "019": ['Turbidity', 'NTU'], 
    "020": ['UVT', '%']
    
}

periodic_table1 = {
    'Ag': 'Silver',
    'Al': 'Aluminium',
    'Au': 'Gold',
    'B': 'Boron',
    'Ba': 'Barium',
    'Be': 'Beryllium',
    'Ca': 'Calcium',
    'Co': 'Cobalt',
    'Cr': 'Chromium',
    'Cu': 'Copper',
    'Fe': 'Iron',
    'K': 'Potassium',
    'La': 'Lanthanum',
    'Mg': 'Magnesium',
    'Mn': 'Manganese',
    'Mo': 'Molybdenum',
    'Na': 'Sodium',
    'Ni': 'Nickel',
    'P': 'Phosphorus',
    'S': 'Sulfur',
    'Sc': 'Scandium',
    'Si': 'Silicon',
    'Sn': 'Tin',
    'Sr': 'Strontium',
    'Ti': 'Titanium',
    'V': 'Vanadium',
    'W': 'Tungsten',
    'Zn': 'Zinc'
}

periodic_table_2 = {
    'As': 'Arsenic',
    'Se': 'Selenium',
    'Cd': 'Cadmium',
    'Sb': 'Antimony',
    'Hg': 'Mercury',
    'Pb': 'Lead',
    'U': 'Uranium'
}

#34 
periodic_table = {
    'Ag': 'Silver',
    'Al': 'Aluminium',
    'Au': 'Gold',
    'As': 'Arsenic', 
    'B': 'Boron',
    'Ba': 'Barium',
    'Be': 'Beryllium',
    'Ca': 'Calcium',
    'Cd': 'Cadmium',
    'Co': 'Cobalt',
    'Cr': 'Chromium',
    'Cu': 'Copper',
    'Fe': 'Iron',
    'Hg': 'Mercury',
    'K': 'Potassium',
    'La': 'Lanthanum',
    'Mg': 'Magnesium',
    'Mn': 'Manganese',
    'Mo': 'Molybdenum',
    'Na': 'Sodium',
    'Ni': 'Nickel',
    'P': 'Phosphorus',
    'Pb': 'Lead',
    'S': 'Sulfur',
    'Sb': 'Antimony',
    'Sc': 'Scandium',
    'Se': 'Selenium',
    'Si': 'Silicon',
    'Sn': 'Tin',
    'Sr': 'Strontium',
    'Ti': 'Titanium',
    'U': 'Uranium',
    'V': 'Vanadium',
    'W': 'Tungsten',
    'Zn': 'Zinc'
}

elementSymbols = {v: k for k, v in periodic_table.items()}
icpMachine2Symbols = ['As', 'Se', 'Cd', 'Hg', 'Pb', 'U']



icpReportRows = ['Ag', 'Al', 'Au', 'B', 'Ba', 'Be', 'Ca', 'Co', 'Cr', 'Cu', 'Fe', 'K', 'La', 'Mg', 'Mn', 'Mo', 'Na', 'Ni', 'P', 'S', 'Sc', 'Si', 'Sn', 'Sr', 'Ti', 'V', 'W', 'Zn', 'As', 'Se', 'Cd', 'Sb', 'Hg', 'Pb', 'U']


GSMS_So_values = {
    "001": 0.100, 
    '002': 0.254, 
    '003': ''
}

GSMS_ref_values = {
    "001": 100, 
    "002": 10, 
    "003": 10, 
        
}


GSMS_std_values ={
    
}



def hardnessCalc(calcium, magnesium): 
    
    return round(float(calcium) * 2.497 + float(magnesium) * 4.11, 1)


def save_pickle(dictonaryName):
    try:
        with open("data.pickle", "wb") as f:
            pickle.dump(dictonaryName, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as ex:
        print("Error during pickling object (Possibly unsupported):", ex)


def load_pickle(filename):
    try:
        with open(filename, "rb") as f:
            return pickle.load(f)
    except Exception as ex:
        print("Error during unpickling object (Possibly unsupported):", ex)
 

def openFile(): 
    fileName, _ = QFileDialog.getOpenFileName(None, 'Open File', '',)
    return fileName

def getFileLocation():
    dlg = QFileDialog().getExistingDirectory()
    print(dlg)
    return dlg

def saveNewLocation():
    location = getFileLocation()
    
    text = input('Save File Name')
    
    save_pickle({text:location})


def scanDir(path): 
    print("Scanning Dir: ", path)    
    
    #print(dir_path)
    obj = os.scandir(path)
    #file = os.listdir(path)
    
    
    for entry in obj :
        if entry.is_dir() or entry.is_file():
            print(entry.name)
    
        


    obj.close()

    

def scanForTXTFolders(jobNum): 
    #print('jobnumber: ', jobNum)

    fileLocationsDictonary = load_pickle('data.pickle')
    TXTLocation = fileLocationsDictonary['TXTDirLocation']
    
    locationsObject = os.scandir(TXTLocation)
    
    txtFolderLocations = [] 
    
    for entry in locationsObject: 
        if(entry.is_dir()):
            if(re.match('^TXT-[a-zA-Z]{3}$', entry.name)):

                txtFolderLocations.append(os.path.join(TXTLocation, entry.name))
            
    
    #print(txtFolderLocations)
    locationsObject.close()
    return processTXTFolders(jobNum, txtFolderLocations)
  
 

def processTXTFolders(jobNum, locations):
    
    fileName = "W" + jobNum + ".TXT"
    
    #print("list")
    #print(locations)
   
    for i in range(len(locations)): 
        tempLocationObject = os.scandir(locations[i]) 

        for entry in tempLocationObject: 
            if(entry.is_file()): 
                if(re.match(fileName, entry.name)): 
                    print("TXT File found")
                    #print(entry.name)
                    tempLocationObject.close()
                    return os.path.join(locations[i], entry.name)
        
        tempLocationObject.close()
    #TODO: return a blank user information 
    #can just clone the clientInfoDict somewhere and send it back 
    print("No Job Number Matches")

        
def processClientInfo(jobNum, fileLocation):
    
    clientInfoDict = {
        'clientName': '', 
        'date': '', 
        'time': '', 
        'attn': '', 
        'addy1': '', 
        'addy2': '', 
        'addy3': '', 
        'sampleType1': '', 
        'sampleType2': '', 
        'totalSamples': '', 
        'recvTemp': '', 
        'tel': '', 
        'email': '', 
        'fax': '', 
        'payment': ''
    }
    
    #grab the file names 
    sampleNames = {}
    

    #have the information about the file, what kind of reports and etc 
    sampleTests = {}

    sampleCounter = 0; 
    prevLine = [0, ""]
    prevLineHelper = [0, ""]
    
    if(fileLocation == None): 
        return clientInfoDict, sampleNames, sampleTests; 
    
    with open(fileLocation) as file: 
    
        for lineLocation, line in enumerate(file, 0):

            if(prevLine[0]+1 == prevLineHelper[0]):
                prevLine[0] = copy(prevLineHelper[0])
                prevLine[1] = copy(prevLineHelper[1])
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
            else: 
                prevLineHelper[0] = copy(int(lineLocation))
                prevLineHelper[1] = copy(line)
            
            #print('PrevLine: ', prevLine[0], prevLine[1])
            #print('PrevLineHelper: ', prevLineHelper[0], prevLineHelper[1])
            #print('currentLine: ', lineLocation, line)
                    
            if(lineLocation == 1): 
                clientInfoDict['clientName'] = line[0:54].strip()
                clientInfoDict['date'] = line[50:(54+7)].strip()
                clientInfoDict['time'] = line[66:71].strip()
                
            if(lineLocation == 2): 
                clientInfoDict['sampleType1'] = line[54:].strip()
                
                if "*" in line: 
                    clientInfoDict['attn'] = line[:54].strip()
                else: 
                    clientInfoDict['addy1'] = line[:54].strip()
                
            if(lineLocation == 3): 
                clientInfoDict['sampleType2'] = line[54:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy1'] = line[:60].strip()
                else: 
                    clientInfoDict['addy2'] = line[:60].strip()
            
            if(lineLocation == 4): 
                clientInfoDict['totalSamples'] = line[60:].strip()
                
                if(clientInfoDict['attn'] != ''):
                    clientInfoDict['addy2'] = line[:60].strip()
                else: 
                    clientInfoDict['addy3'] = line[:60].strip() 
                    
            if(lineLocation == 5): 
                if(clientInfoDict['attn'] and clientInfoDict['addy2']): 
                    clientInfoDict['addy3'] = line[:60].strip()
                else: 
                    clientInfoDict['tel'] = line[26:50].strip()

                    try: 
                        clientInfoDict['recvTemp'] = line[71:].strip()
                    except:
                        print('No recv temp avaliable')
                        
            if(lineLocation == 6): 
                clientInfoDict['tel'] = line[26:50].strip() 
                clientInfoDict['recvTemp'] = line[71:].strip()
            
            if(lineLocation == 7): 
                clientInfoDict['fax'] = line[26:].strip()
                
            if(lineLocation == 8): 
                
                try: 
                    foundEmail = re.search('([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+', line).group()
                    if(foundEmail): 
                        clientInfoDict['email'] = foundEmail; 
                except:
                    print("email error")
                
                if("pd" in line.lower()): 
                    clientInfoDict['payment'] = line[51:].strip()
                     
            
            if(lineLocation > 35 and len(line) > 0): 
               
                if(sampleCounter != int(clientInfoDict['totalSamples']) ):

                    try: 
                        sampleMatch = re.search('(?<=\s[0-9]).*', line).group()
                        if(sampleMatch): 
                            sampleName = str(jobNum) + '-' + str(sampleCounter+1)
                            sampleNames[sampleName] = sampleMatch.strip()
                            sampleCounter+=1; 
                           
                        #TODO: add something to get the - afterwords 
                    except: 
                        pass
                #find the report information that does with corrisponding thing                
                if(re.search('(?<=\s-\s).*', line)):
                    prevSampleName = str(jobNum) + "-" + str(sampleCounter-1)
                    #print('CURRENT: ', line)
                    #print('PREV: ', prevLine[1])
                    currentTestsCheck = re.search('(?<=\s-\s).*', line)
                    prevSampleMatchCheck = re.search('(?<=\s[0-9]).*', prevLine[1])
                    prevSampleTestsCheck = re.search('(?<=\s-\s).*', prevLine[1])
                    #sampleTests[prevSampleName] = currentTestsCheck.group()
                    #not could be apart of the string name longer 
                    
                    #add to most recent sample
                    if(currentTestsCheck):
                        #print('current is a test: ', currentTestsCheck.group())
                        if(prevSampleTestsCheck):
                            sampleTests[prevSampleName] = sampleTests[prevSampleName]  + ", " + currentTestsCheck.group()
                        else: 
                            sampleTests[prevSampleName] = currentTestsCheck.group()
                        
                    #Prev sample name 
                    if(prevSampleMatchCheck):
                        #print('prev was a sampleName')
                        #print(sampleName[prevSampleName]) #doesnt work 
                        pass; 
                       
                    #append onto them 

                        
                    #TODO: solve this later, add previous name onto current name sample 
                    if((not bool(prevSampleMatchCheck)) and not( bool(prevSampleTestsCheck))): 
                        print('prev was apart of the name yo')
                        
                        
            
            #print('---------------------------') 
                
                    
    file.close()
    
    #print(sampleTests)
    #process tyhe sampleTests 
    for key,value in sampleTests.items():
        
        testLists = [x.strip() for x in value.split(',')]
        sampleTests[key] = testLists
           
    
    #print(clientInfoDict)
    #print(sampleNames)
    #print(sampleTests)
    return clientInfoDict, sampleNames, sampleTests; 
    
    
def icp_upload(filePath, db): 
    print('Scanning the file')

    
    if(filePath.endswith('.txt')):
        icpMethod1(filePath, db)
        
    elif(filePath.endswith('.xlsx')):
        icpMethod2(filePath, db)
        
    else: 
        print("Not valid file type")
    
    return; 
    
#TODO: sort by name  
#FIXME: issue sometimes cuts off the ending, so we can just have a cut off section.
#read line by line and just add the line instead 
def icpMethod1(filePath, db): 

    file1 = open(filePath, 'r')
    fname = os.path.basename(filePath)
    #TODO: insert try catch block 
    fname = fname.split('.txt')[0]
    #remove extenion 
    
    print('Method 1')
    print('FileName: ', fname)
    
    Lines = file1.readlines()
    
    startingLine = 'Date Time Label Element Label (nm) Conc %RSD Unadjusted Conc Intensity %RSD' 
    headers = ['Sample', 'Analyte', 'Element', 'HT', ' ', 'units', 'rep', ' ', ' '] 
    
    startingPostion = []
    endPostion = []
    count = 0
    
    # Strips the newline character
    for line in Lines:
        
        #print("Line{}: {}".format(count, line.strip()))
        if(line.strip() == startingLine):
            startingPostion.append(count)
            
        if(re.search('([1-9]|[1-9][0-9]) of ([1-9]|[1-9][0-9])$', line)): 
            endPostion.append(count)

        count += 1

            
    #update headers 
    headerUpdate = Lines[startingPostion[0] + 1]; 
    headerUpdate = headerUpdate.split()
    headers[7] = headerUpdate[0]
    headers[8] = headerUpdate[1]

    newName = fname + '.csv'
    loadPath = load_pickle('data.pickle') 
    print(loadPath)
    newPath = os.path.join(copy(loadPath['ispDataUploadPath']), newName) 
     
    
    print('Writing CSV File: {}'.format(newPath))
    f = open(newPath, 'w')
    writer = csv.writer(f)
    writer.writerow(headers)
    
    spiltLengths = []
    
    jobNumbers = []
    jobData = {
        
    }
    
    elementData = {}
    currentJob = ''
    

    for start in startingPostion: 
        running = True; 
        counter = 1; 

        while(running): 

            currentLine = Lines[start + counter]
            
            if(re.search('([1-9]|[1-9][0-9]) of ([1-9]|[1-9][0-9])$', currentLine)): 
                break; 
            
            counter += 1; 

            splitLine = currentLine.split()
            spiltLengths.append(len(splitLine))

            if(re.search('\d{6}-\d{1,2}', splitLine[2])):
                #create a temp format 
                temp = []
                temp.append(splitLine[2])
                temp.append(1)
                temp.append(splitLine[3])
                temp.append(1)
                temp.append(splitLine[6])
                temp.append('mg/L')
                temp.append(1)
                temp.append(splitLine[0])
                temp.append(splitLine[1])

                if(currentJob =='' ):
                    currentJob = splitLine[2]
  
                elif(currentJob != splitLine[2]):
                    jobData[currentJob] = elementData   
                    elementData = {}
                    currentJob = splitLine[2]

                elementData[splitLine[3]] = splitLine[6]
    
                jobNumber = splitLine[2].split('-')[0]
                if(jobNumber not in jobNumbers):
                    #print(jobNumber)
                    jobNumbers.append(jobNumber)
                    
                if(temp): 
                    writer.writerow(temp)
            
            #print(len(splitLine))

    
    spiltLengths = numpy.array(spiltLengths)
    unique, counts = numpy.unique(spiltLengths, return_counts=True)
    #print(dict(zip(unique, counts)))

    f.close()
    file1.close()
    
    #print(jobNumbers)
    
    #FIXME: uploading same data
    #we can have a check in place in the file where it is saved 
    
    #save to database
    todayDate = date.today()
    #save to database 
    for (key, value) in jobData.items(): 
        #print(key)
        sql = 'INSERT INTO icpMachineData1 values(?,?,?,?,?, 1)'
        jobNum = key.split('-')[0]
        tempData = json.dumps(value)
        db.execute(sql, (key,jobNum,newPath, tempData, todayDate))
        db.commit()
    
    
    return jobNumbers, jobData; 

#scans throught all the text files and finds all the different sample types and the ISP and GSMS files     
def icpMethod2(filePath, db): 
    
    wb = openpyxl.load_workbook(filePath)
    sheets = wb.sheetnames 
    
    fname = os.path.basename(filePath)
    #TODO: insert try catch block 
    fname = fname.split('.xlsx')[0]
    
    print('Method 2')
    print('FileName: ', fname)


    newName = fname + '.csv'
    loadPath = load_pickle('data.pickle') 
    
    newPath = os.path.join(copy(loadPath['ispDataUploadPath']), newName) 
    
    worksheet = wb[sheets[0]]

    sampleTypeColumn = worksheet['E']
    sampleNameColumn = worksheet['G']
    
    elementConversion = ['He', 'Se', 'Cd', 'Sb', 'Hg', 'Pb', 'U']
    elementColumns = ['I', 'J', 'M', 'O', "Q", 'U', 'AA', 'AC']
    selectedRows = []
    
    
    for cell in sampleTypeColumn:     
        if(cell.value == 'Sample'):
            pass 
            #print(cell.value)
            currentSampleName = worksheet.cell(row=cell.row, column=7).value
           
            if any([x in currentSampleName for x in ['blk', 'curve']]):
                
                continue; 
            else: 
                print(currentSampleName)
                selectedRows.append(cell.row)
            #    print(f'Row: {cell.row} Value: {currentSampleName}')
             
    #    print(f'Row: {cell.row} value: {cell.value}')    
    
    #print(selectedRows)
    
    for row in selectedRows: 
        row_values = []
        for col in elementColumns:
            col_index = openpyxl.utils.column_index_from_string(col)
            cell_value = worksheet.cell(row=row, column=col_index).value
            row_values.append(cell_value)
        #print(row_values)
    
    #write excel 
    
    newWB = openpyxl.Workbook()
    worksheet2 = newWB.active 
    
    worksheet2.cell(row=1, column=1).value = 'Sample'
    tableNames = ['', 'Rjct', 'Data File', 'Acq. Date-Time', 'Type', 'Level', 'Sample Name', 'Total Dil.']
    
    worksheet2.merge_cells('A1:H1')
    
    column_num = 1;
    for item in tableNames: 
        worksheet2.cell(row=2, column=column_num).value = item; 
        column_num += 1; 
    
   
    row_num = 3; 
    for row_value in selectedRows:

        for currentPos in range(1,8): 
            worksheet2.cell(row=row_num, column=currentPos).value = worksheet.cell(row=row_value, column=currentPos).value
        
        tempCol = 9
        for col in elementColumns: 
            col_index = openpyxl.utils.column_index_from_string(col)
            worksheet2.cell(row=row_num, column=tempCol).value = worksheet.cell(row=row_value, column=col_index).value
            tempCol+=1; 
            

        row_num+=1;      
        
 
     
    
    newWB.save('text.xlsx')

        
        
    
    return; 



#FIXME: if another thing comes along 
def createReport(db, jobNum, reportType, parameter=False, recovery=False): 
    
    if('ISP' in reportType):
        sql  = 'INSERT INTO icpReports (jobNumber, reportType, status, createdDate) values (?,?,1,?)'
        
    if(reportType == 'GSMS'):
        sql  = 'INSERT INTO gsmsReports (jobNumber, reportType, status, createdDate) values (?,?,1,?)'

    todayDate = date.today()
    print('COMMAND: ', sql) 

    db.execute(sql, (jobNum, reportType, todayDate))
    db.commit()
    

    
def saveClientInformation(db): 
    
    pass; 

def saveGSMSData(): 
    
    
    pass; 