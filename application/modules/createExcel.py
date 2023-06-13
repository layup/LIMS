
import os 

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, borders, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText 
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image


from openpyxl.worksheet.header_footer import HeaderFooter

import math 
import re 
import os 

from modules.utilities import * 


pageRows = 46
defaultFont = Font(name="Times New Roman", size=9)

#Borders 
thinBorder = Side(border_style="thin", color="000000")
doubleBorder = Side(border_style='double', color="000000")



def createGcmsReport(clientInfo, jobNum, sampleNames, sampleData, testInfo, unitType, recovery):
    print('CREATING GCMS REPORT')
    print(clientInfo)
    print(sampleNames)
    print(sampleData)
    print(testInfo)
    print(unitType)
    print(recovery)
    print("--------------------")
   
    #FIXME: picke saving and loading without being fucked up, get path of where things should be saved
    temp = load_pickle('data.pickle') 
    exportPath = temp['reportsPath']
    
    wb = Workbook()
    ws = wb.active
    
    pageSetup(ws); 
    createFooters(ws); 

    #SETTING UP MAIN INFORMATION 
    #TODO: determine how far out the longest thing is 
    #FIXME: check out for none issue when they appear
    
    ws = createHeader(ws, clientInfo, 'D')
    ws.column_dimensions['A'].width = 20 #120px 
    #ws.column_dimensions['G'].width = 14 
    ws.column_dimensions['H'].width = 19
    ws.print_title_rows = '1:8' # the first two rows
    
    
    #INSERT NEW COLUMN FOR So
    totalCols = 8 
    pageSize = 61 
    
    
    #FIXME: add the proper job number instead of placeholder 
    jobNumLocation = ws.cell(row=1, column=8)
    jobNumLocation.value = "W" + str(jobNum)
    jobNumLocation.alignment = Alignment(horizontal='right', vertical='center')

    sampleSections = []
    samplePlacement = []
    currentWord = ''
    temp = []
    
    totalSamples = len(sampleData.keys())
    print('Total Samples: ', totalSamples)
    #FIXME: so page amounts change based on column width, so takes in how many pages we want 
    formatRows(ws, totalSamples, totalCols, pageSize);
    
    #can only display 4 at a time 
    currentWord = ""
    temp = []
    
    for i, (key,value) in enumerate(sampleNames.items(), start=1): 
        
        stripedWord = " ".join(value.split())

        #print(key," ".join(value.split()))
        currentWord += " " + str(i) + ") " + stripedWord + " "
        temp.append(key)

        if(i % 4 == 0): 
            print(currentWord)
            sampleSections.append(currentWord)
            samplePlacement.append(temp)
            currentWord = ""
            temp = []
    
    if(currentWord != ''):
        sampleSections.append(currentWord)
        samplePlacement.append(temp) 
    
    print('Sample Placement: ', samplePlacement)
    print('Sample Sections: ', sampleSections)
    
    
    pageSize = 56; 
    allocatedSpace = 20; 
    
    testSize = len(testInfo)
    tableSize = 6 + testSize 
    totalSampleSections = len(sampleSections)
    print('Total Sample Sections: ', totalSampleSections)
    
    totalTablesWithComments = math.floor((pageSize - allocatedSpace)/tableSize)
    print('Tables with comments: ' , totalTablesWithComments)
    
    #totalTablesWithoutComments = math.floor((pageSize - (8+6))/tableSize)
    #print('Tables without comments: ', totalTablesWithoutComments)
    
    totalPages = math.ceil(totalSampleSections/totalTablesWithComments)
    print('Total Pages: ', totalPages)
    
    usedSamples = 0; 
    pageLocation = 9; 
    totalTests = len(testInfo)

    #determine when the next page is? 
    counter = 0; 
    
    for currentPage in range(totalPages): 
        
        print('current page:', currentPage)
        sampleAmount = len(samplePlacement[counter])
  
        
        if(currentPage != 0): 
            pageLocation = (pageSize * currentPage) - (8 * (currentPage-1)) + 1  
        else: 
            pageLocation = 9; 
        
        #last page 
        if(currentPage+1 == totalPages): 
 
            remainingSamples = totalSampleSections - counter; 
            
            for i in range(remainingSamples): 
                pageLocation = insertSampleName(ws, pageLocation, sampleSections[counter], totalCols)
                pageLocation = insertTestTitles(ws,pageLocation, sampleAmount, usedSamples, 0, totalCols)
                pageLocation = insertTestInfo(ws,pageLocation, testInfo, samplePlacement[counter], sampleData, totalTests, unitType, recovery)
                
                if(i+1 == remainingSamples): 
                    pageLocation = insertComments(ws, pageLocation)  
                    pageLocation+=1; 
             
                    img = Image('signature.png')
                    cell = ws.cell(row=pageLocation, column=3)
                    coordinate = cell.coordinate
                    ws.add_image(img, coordinate)
                        
                counter+=1; 
                usedSamples += 4;
        
        #not the last page
        else: 
            print('Not Last Page')
            for i in range(totalTablesWithComments): 
                pageLocation = insertSampleName(ws, pageLocation, sampleSections[counter], totalCols)
                pageLocation = insertTestTitles(ws,pageLocation, sampleAmount, usedSamples, 0, totalCols)
                pageLocation = insertTestInfo(ws,pageLocation, testInfo, samplePlacement[counter], sampleData, totalTests, unitType, recovery)   
                
                if(i+1 == totalTablesWithComments): 
                    comment = ws.cell(row=pageLocation, column=1)
                    comment.value = 'continued on next page....'
                    comment.font = Font(bold=True, size=9, name="Times New Roman")  
                    
                counter+=1; 
                usedSamples += 4; 
        
    
    fileName = 'W' + str(jobNum) + ".chm" 
    filePath = os.path.join(exportPath, fileName)
    print('Export Path: ', filePath)
    
    wb.save(filePath)
    

def createIcpReport(clientInfo, sampleNames, jobNum,  sampleData, testInfo, unitType, limitElements, limits, foooterComment): 
    print('ICP general Information')
    print(clientInfo)
    print(sampleNames)
    print(sampleData)
    print(testInfo)
    print(unitType)
    print(limitElements)
    print(limits)
    print('-------------------------')
    #TODO: missing the report type that will get called into this thing 
    temp = load_pickle('data.pickle') 
    exportPath = temp['reportsPath']

    newList = [item.lower() for item in testInfo]
    
    print(newList)
    limitRef = {}
    
    for i, item in enumerate(limitElements): 
        try: 
            index = newList.index(item)
            limitRef[index] = limits[i]
        except:
            print("could not find") 
            
    print(limitRef)

    wb = Workbook()
    ws = wb.active
    
    pageSetup(ws);    
    createFooters(ws); 

    #SETTING UP MAIN INFORMATION 
    #TODO: determine how far out the longest thing is 
    #FIXME: check out for none issue when they appear
    #TODO: add the job number to the end of this thing mf 
        
    ws = createHeader(ws, clientInfo, 'D')
    ws.column_dimensions['A'].width = 20 #120px 
    ws.column_dimensions['H'].width = 19 
    ws.print_title_rows = '1:8' # the first two rows

    #FIXME: add this to the header information  
    jobNumLocation = ws.cell(row=1, column=8)
    jobNumLocation.value = "W" + str(jobNum)
    jobNumLocation.alignment = Alignment(horizontal='right', vertical='center')
    
    #determine how many rows the sample name will take 
    sampleSections = []
    samplePlacement = []    
    selectedNames = []
    
    currentWord = ''
    temp = []
    
    #determine how many sample sections we will need 
    for key in sampleData: 
        selectedNames.append(key)
    
    #can only display 4 at a time 
    print('!------ GENERATING SAMPLE NAME ------!')
    for i, sampleNum in enumerate(selectedNames, start=1): 
        sampleName = sampleNames[sampleNum]
        condencedName = " ".join(sampleName.split())
        currentWord += " " + str(i) + ") " + condencedName + " "
        temp.append(sampleNum)

        if(i % 4 == 0): 
            sampleSections.append(currentWord)
            samplePlacement.append(temp)
            currentWord = ""
            temp = []
        
    if(len(temp) != 0): 
        sampleSections.append(currentWord)
        samplePlacement.append(temp)
        currentWord = ""
        temp = [] 
            

    #insert sample names section
    print('Sample Placement: ', samplePlacement)
    print('Sample Sections: ', sampleSections)
    totalTests = len(testInfo)
    totalPages = len(samplePlacement)
    
    
    #for loop that goes through each page, tell the starting page 
    #insert the sample names 
    #insert the test headers 
    #insert the test data 
    #insert the ending or continue to next page  
    
    #793 = total size of page 1 
    #first page has 61 pages (61 + 61 - 8) or (61 * (total) - (8 * total - 1))
    #next page is only 45 pages + 8 (not accounted for)
    #start = 9, insert 41 things, end = 50, page end = 61 
    #[0,1,2]
    #if 0, start at 9
    #if 1 = (61 * i) - (8 * (i-1)) + 1 =  62
    #if 2 = (61 * 2) - (8 * 1) + 1 = 115 
    #if 2 = (61 * 3) - (8 * 2) + 1 = 
    
    
    #TODO: tests how many of the tests need footer comments 
    #TODO: determine how long the extra tests comments are 
    #TODO: factor that into the size insertion 
    totalCols = 8 
    pageSize = 61; 
    pageLocation = 9; 
    usedSamples = 0; 
    
    #FIXME: include the length of the comments 
    totalSamples = len(sampleData.keys())
    
    print('Total Samples: ', totalSamples);
    formatRows(ws, totalSamples, totalCols, pageSize)
    #how do we know where to start each loop? 
    for currentPage in range(totalPages): 
        
        sampleAmount = len(samplePlacement[currentPage])
        print('Sample Amount: ', sampleAmount);
        
        #first page infomation 
        if(currentPage == 0): 
            pageLocation = 9; 
            pageLocation = insertSampleName(ws, pageLocation, sampleSections[currentPage], totalCols)
            pageLocation = insertTestTitles(ws, pageLocation, sampleAmount, usedSamples, 1, totalCols ) 
            pageLocation = insertIcpTests(ws, pageLocation, totalTests, testInfo, unitType, samplePlacement[currentPage], sampleData, limitRef) 
            
            #TODO: insert the comment 
            if(totalPages > 1): 
                comment = ws.cell(row=pageLocation, column=1)
                comment.value = 'Continued on next page ....'
                comment.font = Font(bold=True, size=9, name="Times New Roman")
            else: 
                print('Insert comment and signature') 
                
                for (i, value) in enumerate(foooterComment):
                    comment = ws.cell(row=pageLocation, column=1)
                    comment.value = value
                    comment.font = Font(size=9, name="Times New Roman") 
                    pageLocation+=1; 
                    
                pageLocation+=1; 
                img = Image('signature.png')
                cell = ws.cell(row=pageLocation, column=4)
                coordinate = cell.coordinate
                ws.add_image(img, coordinate)


            
        else: 
            pageLocation = (61 * currentPage) - (8 * (currentPage-1)) + 1 
            print('Starting Location: ', pageLocation) 
            pageLocation = insertSampleName(ws, pageLocation, sampleSections[currentPage], totalCols)
            pageLocation = insertTestTitles(ws, pageLocation, sampleAmount, usedSamples, 1, totalCols) 
            pageLocation = insertIcpTests(ws, pageLocation, totalTests, testInfo, unitType, samplePlacement[currentPage], sampleData, limitRef) 
            
            if((currentPage+1) == totalPages): 
                for (i, value) in enumerate(foooterComment):
                    comment = ws.cell(row=pageLocation, column=1)
                    comment.value = value
                    comment.font = Font(size=9, name="Times New Roman") 
                    pageLocation+=1; 


                pageLocation+=1; 
                img = Image('signature.png')
                cell = ws.cell(row=pageLocation, column=4)
                coordinate = cell.coordinate
                ws.add_image(img, coordinate)
            else: 
                comment = ws.cell(row=pageLocation, column=1)
                comment.value = 'Continued on next page ....'
                comment.font = Font(bold=True, size=9, name="Times New Roman")
        
        usedSamples += sampleAmount; 
    
    maxWidth = ws.max_column 
    print(f'The width of the worksheet is {maxWidth} columns')
    print('Current Page location: ', pageLocation); 
    
     
    fileName = 'W' + str(jobNum) + ".001" 
    filePath = os.path.join(exportPath, fileName)
    print('Export Path: ', filePath)
 
    wb.save(filePath)

def pageSetup(ws): 
   
    # set the default view to page layout
    ws.sheet_view.view = "pageLayout" 

    # Set the page width to auto
    page_setup = ws.page_setup
    
    #setup page size 
    page_setup.fitToPage = True
    page_setup.fitToHeight = False 
    page_setup.fitToWidth = True
    
    #page margins     
    page_margins = PageMargins()
    page_margins.left = 0.7
    page_margins.right = 0.7
    page_margins.top = 0.75
    page_margins.bottom = 0.75
    page_margins.header = 0.3
    page_margins.footer = 0.3
    
    ws.page_margins = page_margins
    
#FORMATS WAY MORE THEN NEEDED BUT W/
def formatRows(ws, totalSamples, totalCols, pageSize): 

    totalPages = math.ceil(totalSamples/4) 
    print('Total Pages: ', totalPages)

    totalRows = (pageSize * totalPages) - (8 * (totalPages-1))
    print('Total Rows: ', totalRows)

    for row in ws.iter_rows(min_row=1, max_col=totalCols, max_row=totalRows): 
        
        for cell in row:
            cell.font = defaultFont 
            
            #if(cell.row < 10): 
            ws.row_dimensions[cell.row].height = 13
            


def createFooters(ws): 
    
    #Setting up the headers and footers 
    ws.oddHeader.fontName = 'Times New Roman'
    ws.oddHeader.fontSize = 14

    ws.evenHeader.left.font_name = 'Times New Roman'
    ws.evenHeader.left.font_size = 14
    
    ws.oddHeader.left.text  = 'GCMS - Report Form: &D'
    ws.evenHeader.left.text = 'GCMS - Report Form: &D' 
    
    
    ws.oddHeader.right.text  = "Page &P of &N"
    ws.evenFooter.right.text = "Page &P of &N"
    
    ws.oddFooter.left.text  = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com'
    ws.evenFooter.left.text = '&BT:&B 250 656 1334 \n&BE:&B info@mblabs.com' 
    
    ws.oddFooter.center.text = "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    ws.evenFooter.center.text= "&B MB Laboratories Ltd.&B \nwww.mblabs.com "
    
    ws.oddFooter.right.text = '&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    ws.evenFooter.right.text ='&BMail:&B PO BOX 2103 Stn Main \n Sidney, B.C, V8L 356'
    

def createHeader(ws, clientInfo, column2): 
    
    ws['A1'] = clientInfo['clientName']
    
    #font = Font(size=9)
    #ws['A1'].font = font
    
    if(clientInfo['attn'] is None): 
        ws['A2'] = '*' 
    else: 
        ws['A2'] = clientInfo['attn'] 
        
    ws['A3'] = clientInfo['addy1']
    ws['A4'] = clientInfo['addy2'] + ", " + clientInfo['addy3']
    
    ws['A6'] = 'TEL' + clientInfo['tel']
    ws['A7'] =  clientInfo['email']
    
    ws[column2 + '1'] = "Date: " + clientInfo['date'] + "  (" + clientInfo['time'] + ")" 
    ws[column2 + '2'] = "Source: " + clientInfo['sampleType1']
    ws[column2 + '3'] = "Type: " + clientInfo['sampleType2']
    ws[column2 + '4'] = "No. of Samples: " + clientInfo['totalSamples']
    ws[column2 + '5'] = "Arrival temp: " + clientInfo['recvTemp']
    ws[column2 + '6'] = "PD: " + clientInfo['payment']
    
    return ws 
     
def createComments(ws):
    pass;      

def mergeSampleSection(ws, row, col): 
    
    pass; 

def nextPage(curVal, curPage): 
    
    #minue 9 for each interation that we increase 
    
    pageEnds = [47]
    
    for i in range(10):
        pageEnds.append(pageEnds[i] + 37)
        
    
    if(curVal > pageEnds[curPage]): 
        print('FALSE')
        return False 
    else: 
        print("TRUE")
        return True; 

def insertSampleName(ws, row, sampleSection, totalRows): 
    
    temp = ws.cell(row=row, column=1)
    temp.value = 'Samples: ' + sampleSection
    temp.border = Border(bottom=thinBorder)
   
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=totalRows)
    temp.alignment = Alignment(wrap_text=True) 
    
    return row + 3; 
    
def insertTestTitles(ws, pageLocation, totalSamples, startVal, reportType, totalCols): 
    
    tests = ws.cell(row=pageLocation, column=1)
    if(reportType == 0):
        tests.value = 'Tests'
    else: 
        tests.value = 'Elements'        

    units = ws.cell(row=pageLocation, column=2)
    if(reportType == 0): 
        units.value = 'Units'
    else: 
        units.value = "Symbols"

    units.border = Border(right=thinBorder, left=thinBorder)
    units.alignment = Alignment(horizontal='center', vertical='center')
    
    allowedBorders = [1,2,7,8]
    
    #assign sample names 
    for i in range(totalSamples): 
        sample = ws.cell(row=pageLocation, column=i+3)
        allowedBorders.append(i+3)
        
        if(startVal == 0): 
            sample.value = 'Sample ' + str(i + 1)
        else: 
            sample.value = 'Sample ' + str(startVal + i +1) 
             
        sample.alignment = Alignment(horizontal='center', vertical='center')
        sample.border = Border(right=thinBorder, left=thinBorder) 


    so = ws.cell(row=pageLocation, column=7) 
    if(reportType == 0): 
        so.value = 'Recovery'
    else: 
        so.value = 'Units'
    so.border = Border(right=thinBorder, left=thinBorder)
    so.alignment = Alignment(horizontal='center', vertical='center') 
    
    
    colH = ws.cell(row=pageLocation, column=8)
    if(reportType == 0 ): 
        colH.value = 'Comment'
        colH.alignment = Alignment(horizontal='center', vertical='center')  
        
    else:
        colH.value = 'Maximum Limits'  
        colH.alignment = Alignment(horizontal='left', vertical='center', indent=1)  
    
    pageLocation+=1; 
    
        
    for i in range(1,totalCols+1): 
        current = ws.cell(row=pageLocation, column=i) 
        
        if(i in allowedBorders): 
        
            if(i != 8): 
                current.border = Border(right=thinBorder, bottom=doubleBorder)
            else: 
                current.border = Border(bottom=doubleBorder) 
            
            if(i == 7): 
                current.border = Border(right=thinBorder, left=thinBorder, bottom=doubleBorder)
            
        else: 
            current.border = Border(bottom=doubleBorder)
            
         
    pageLocation+=1; 

    return pageLocation 


#will change how we get the tests and unit type, will be saved somewhere 
def insertTestInfo(ws, pageLocation, testInfo, samplePlacement, sampleData, totalTests, unitType, recovery): 

    counter = pageLocation

    for i in range(len(testInfo)): 
        #is this removing the escape characters? 
        #currentTest = re.sub('[^A-Za-z0-9]+', '', testInfo[i])
        testPlacement = ws.cell(row=counter, column=1)
        unitPlacement = ws.cell(row=counter, column=2)
        recoveryPlacement = ws.cell(row=counter, column=7)

        try:  
            testPlacement.value = testInfo[i]
        except: 
            testPlacement.value = 'Error'
            
        unitPlacement.value = unitType[i]
        recoveryPlacement.value = recovery[i]
        
        #testPlacement.border = Border(right=thinBorder) 
        unitPlacement.border = Border(right=thinBorder, left=thinBorder)
        unitPlacement.alignment = Alignment(horizontal='center', vertical='center')
        
        recoveryPlacement.border = Border(left=thinBorder, right=thinBorder)
        recoveryPlacement.alignment = Alignment(horizontal='center', vertical='center')  
        

        
        counter+=1; 
    
    
    counter = pageLocation 
    
    #insert test user information
    for i, sample in enumerate(samplePlacement, start=3): 
        
        currentResults = sampleData[sample]
        
        for j in range(0,totalTests): 
            currentSample = ws.cell(row=counter+j,column=i)
            currentSample.alignment = Alignment(horizontal='center', vertical='center')
            
            if(type(currentResults[j]) == (int or float)): 
               
                currentSample.value = float(currentResults[j])
            else:
                currentSample.value = currentResults[j]
            
            currentSample.value = currentResults[j]
            currentSample.border = Border(right=thinBorder, left=thinBorder)
                     
    pageLocation += totalTests; 
    
    for i in range(1,9): 
        ws.cell(row=pageLocation, column=i).border = Border(top=thinBorder)
            
    pageLocation += 1; 

    
    return pageLocation; 

def insertComments(ws, pageLocation): 

    comments = { 
        "SD":  'SD    = standard devition;       Standard Recovery = primary or secondary reference material', 
        'STD': 'STD  = secondary standard calibrated to primary standard reference material', 
        'ND':  'ND   = none is detcted;          n/a = not applicable'
    }
    
    for i, (key,value) in enumerate(comments.items()): 
        temp = ws.cell(row = pageLocation+i, column=1)
        temp.value = value
        
    pageLocation+=5; 
    return pageLocation; 



def insertSignature(ws, pageLocation): 
    
    name1 = 'R. Bilodeau'
    title1 = 'Analytical Chemist'
    name2 = 'H. Hartmann'
    

def insertIcpTests(ws, pageLocation, totalTests, testInfo, unitType, samplePlacement, sampleData, limitRef): 
    #variables 

    counter = pageLocation; 
    
    #insert elementName, symbol 
    #FIXME: need to change because will be loading elements and symbols based on the info 
    for item in range(totalTests): 
        elementRow = ws.cell(row=counter, column=1); 
        symbolRow = ws.cell(row=counter, column=2); 
        unitRow = ws.cell(row=counter, column=7)
        
        elementRow.value = "{0}) {1}".format(item+1, testInfo[item])

        temp = testInfo[item]
        symbolRow.value = elementSymbols[temp]
        #try: 
        #    unitRow.value = unitType[item]
        #except: 
        #    print("error")
        
        elementRow.border = Border(right=thinBorder) 
        symbolRow.border  = Border(right=thinBorder) 
        unitRow.border    = Border(right=thinBorder, left=thinBorder) 
        
        symbolRow.alignment = Alignment(horizontal='center', vertical='center')
        unitRow.alignment   = Alignment(horizontal='center', vertical='center')
        
        counter+=1; 
   
    counter = pageLocation 
    sampleLocation = 0 

    #inserts the sample values, unitType, limits and comments 
    for currentCol, sample in enumerate(samplePlacement, start=3): 
        
        print(currentCol, sample)
        
        for j in range(0, totalTests+2): 
            currentSample = ws.cell(row=counter+j, column=currentCol)        
            currentSample.alignment = Alignment(horizontal='center', vertical='center')  
            currentSample.border = Border(right=thinBorder)
            
            currentUnit = ws.cell(row=counter+j, column=7); 
            
            comment = ws.cell(row=counter+j, column=8)
            comment.alignment = Alignment(horizontal='left', vertical='center', indent=1)   
            
            currentVal = sampleData[sample][j]
            #print(j, currentVal)
            #print(j)
            
            #FIXME: this is wrong atm uncal means something else 
            try: 
                temp = float(currentVal)
        
                if j in limitRef: 
                    lower = limitRef[j][1]
                    higher = limitRef[j][2]
                    currentSample.value = temp 
                    #print('{} Lower: {} | Highser: {}'.format(j, lower,higher))
                    
                    if(lower != ''): 
                        if(temp < lower): 

                            currentSample.value = '< ' + f'{lower:.3f}'

                    if(higher != ''): 
                        if(temp > higher): 
                            currentSample.value = '> ' + f'{higher:.3f}'
                    
                    if(temp == 0): 
                       currentSample.value = 'ND' 
                
                #no limit exists for the given thing
                else: 
                    currentSample.value = temp 
                    
                    if(temp == 0): 
                        currentSample.value = 'ND'  

            except:
                if(currentVal == 'Uncal'):
                    currentSample.value = 'n/a'
                else: 
                    currentSample.value = currentVal
                    
            if j in limitRef: 
                higher = limitRef[j][2]
                limitComment = limitRef[j][3]
                unitType = limitRef[j][4]

                currentUnit.value = unitType

                if(limitComment != ''): 
                    comment.value = limitComment
                #FIXME: insert the unit value if exists 
                elif(higher != ''): 
                    if(higher < 1): 
                        comment.value = f'{higher:.3f}' + " " + unitType   
                    else: 
                        comment.value = f'{higher:.2f}' + " " + unitType
                    
                    if(higher > 10):  
                        comment.value = f'{higher:.1f}' + " " + unitType  
                else: 
                    comment.value = 'no limit listed'
            else: 
                comment.value = 'no limit listed' 
                
            #print(currentSample.value) 
    
    
    sampleLocation += 1 
    pageLocation += totalTests;  

    additionalTestsColumns = [1,2,7,8]
    
    for current in additionalTestsColumns: 
        
        temp = ws.cell(row=pageLocation, column=current)
        
        if(current != 8): 
            temp.border = Border(right=thinBorder)
        else: 
            temp.value = '0-75 mg/L = soft'
            temp.alignment = Alignment(horizontal='left', vertical='center', indent=1)   
        
        if(current == 1):
            temp.value = 'Hardness'
            
        if(current == 2): 
            temp.value = 'CaCO₃'
            temp.alignment = Alignment(horizontal='left', vertical='center', indent=1)   
            
        if(current == 7):
            temp.value = 'mg/L'
            temp.alignment = Alignment(horizontal='center', vertical='center') 
            temp.border = Border(right=thinBorder, left=thinBorder)
            
    pageLocation+=1; 

    for current in additionalTestsColumns: 
        temp = ws.cell(row=pageLocation, column=current)
       
        if(current != 8): 
            temp.border = Border(right=thinBorder)
        else: 
            temp.value = '7.0 to 10.5' 
            temp.alignment = Alignment(horizontal='left', vertical='center', indent=1)   
        
        if(current == 1):
            temp.value = 'Ph'
        
        if(current == 7):
            temp.value = 'units'
            temp.alignment = Alignment(horizontal='center', vertical='center') 
            temp.border = Border(right=thinBorder, left=thinBorder)
            
    pageLocation+=1; 
    
    for i in range(1,9): 
        bottomBorder = ws.cell(row=pageLocation, column=i)
        bottomBorder.border = Border(top=thinBorder)
            
    pageLocation +=1; 
    
    return pageLocation; 